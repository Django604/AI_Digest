from __future__ import annotations

import calendar
import json
from datetime import date
from pathlib import Path
import shutil
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from scripts.build_dashboard import (
    ARRIVAL_BOOK,
    LEADS_BOOK,
    MONTHLY_ARCHIVE_DIR,
    NEV_CORE_MODELS,
    NEV_DETAIL_MODELS,
    NEW_PATHFINDER_TARGET_OVERRIDES,
    SYLPHY_FREEZE_DATE,
    OUT_JSON,
    SUMMARY_JSON,
    apply_submission_time,
    apply_preserved_input_modified_times,
    build_arrival_dashboard,
    build_arrival_series,
    build_arrival_brief,
    build_column_meta,
    build_payload,
    build_run_summary,
    build_nev_section,
    build_valid_leads_brief,
    build_valid_leads_control_trend,
    file_mtime_iso,
    get_day_calendar_meta,
    load_preserved_input_modified_times,
    load_arrival_daily_sheet,
    load_valid_leads_monthly_targets,
    resolve_new_pathfinder_targets,
    validate_report_date_cell,
    validate_sheet_headers,
    validate_workbook_structure,
    write_monthly_archive,
    write_json_if_changed,
)


class BuildDashboardPayloadTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.payload = build_payload(LEADS_BOOK, ARRIVAL_BOOK)
        report_date = date.fromisoformat(cls.payload["meta"]["reportDate"])
        previous_year = report_date.year if report_date.month > 1 else report_date.year - 1
        previous_month = report_date.month - 1 if report_date.month > 1 else 12
        previous_month_end = date(
            previous_year,
            previous_month,
            calendar.monthrange(previous_year, previous_month)[1],
        )
        cls.previous_month_payload = build_payload(
            LEADS_BOOK,
            ARRIVAL_BOOK,
            report_date_override=previous_month_end,
        )
        cls.previous_month_key = f"{previous_month_end.year:04d}-{previous_month_end.month:02d}"
        cls.previous_month_archive_path = MONTHLY_ARCHIVE_DIR / cls.previous_month_key / "dashboard.json"
        cls.previous_month_archive_payload = (
            json.loads(cls.previous_month_archive_path.read_text(encoding="utf-8"))
            if cls.previous_month_archive_path.exists()
            else None
        )
        cls.synthetic_date = date(2026, 7, 10)
        cls.synthetic_report_date = date(2026, 7, 13)
        cls.synthetic_nev_daily = {
            "NX8": {
                cls.synthetic_date: {
                    "newLeads": 10,
                    "validLeads": 7,
                    "storeLeads": 8,
                    "arrivals": 1,
                }
            },
            "2026款探陆": {
                cls.synthetic_date: {
                    "newLeads": 2,
                    "validLeads": 1,
                    "storeLeads": 1,
                    "arrivals": 0,
                }
            },
        }
        synthetic_targets = {"NX8": {cls.synthetic_date: 20}}
        with (
            patch("scripts.build_dashboard.load_nev_daily", return_value=cls.synthetic_nev_daily),
            patch("scripts.build_dashboard.load_nev_targets", return_value=synthetic_targets),
            patch("scripts.build_dashboard.NEW_PATHFINDER_TARGET_OVERRIDES", {}),
        ):
            cls.synthetic_payload = build_payload(
                LEADS_BOOK,
                ARRIVAL_BOOK,
                report_date_override=cls.synthetic_report_date,
            )
        synthetic_without_new_pathfinder = {
            model_name: series
            for model_name, series in cls.synthetic_nev_daily.items()
            if model_name != "2026款探陆"
        }
        with (
            patch("scripts.build_dashboard.load_nev_daily", return_value=synthetic_without_new_pathfinder),
            patch("scripts.build_dashboard.load_nev_targets", return_value=synthetic_targets),
            patch("scripts.build_dashboard.NEW_PATHFINDER_TARGET_OVERRIDES", {}),
        ):
            cls.synthetic_payload_without_new_pathfinder = build_payload(
                LEADS_BOOK,
                ARRIVAL_BOOK,
                report_date_override=cls.synthetic_report_date,
            )

    def test_expected_dashboards_exist(self) -> None:
        self.assertEqual(
            set(self.payload["dashboards"].keys()),
            {"brief", "lead-control", "nev", "ice", "arrival"},
        )

    def test_nev_model_groups_keep_new_pathfinder_out_of_core_total(self) -> None:
        core_model_names = [model_name for _, _, model_name in NEV_CORE_MODELS]
        detail_model_names = [model_name for _, _, model_name in NEV_DETAIL_MODELS]

        self.assertEqual(core_model_names, ["NX8", "N7", "N6", "天籁·鸿蒙座舱"])
        self.assertEqual(detail_model_names, [*core_model_names, "2026款探陆"])

    def test_nev_total_chart_title_identifies_four_model_scope(self) -> None:
        section = next(
            item for item in self.synthetic_payload["dashboards"]["nev"]["sections"]
            if item["id"] == "nev-total"
        )

        self.assertEqual(section["trend"]["chartTitle"], "7 月NEV 4车新增线索趋势")

    def test_sylphy_history_is_preserved_but_not_rendered(self) -> None:
        ice_section_ids = [item["id"] for item in self.payload["dashboards"]["ice"]["sections"]]
        brief_kinds = [
            item["kind"]
            for item in self.payload["dashboards"]["brief"]["briefing"]["sections"]
        ]

        self.assertIn("十五代轩逸按日", self.payload["analysis"]["sheetNames"])
        self.assertNotIn("sylphy-15", ice_section_ids)
        self.assertNotIn("sylphy15", brief_kinds)

    def test_new_pathfinder_section_follows_existing_nev_models(self) -> None:
        sections = self.payload["dashboards"]["nev"]["sections"]

        self.assertEqual([section["title"] for section in sections][-2:], ["天籁·鸿蒙座舱", "2026款探陆"])
        self.assertEqual(sections[-1]["id"], "new-pathfinder")

    def test_new_pathfinder_section_removes_dates_before_july_16(self) -> None:
        section = next(
            item for item in self.synthetic_payload["dashboards"]["nev"]["sections"]
            if item["id"] == "new-pathfinder"
        )
        matrix = section["trend"]["matrix"]
        self.assertEqual(matrix["labels"], [f"7/{day}" for day in range(16, 32)])
        self.assertNotIn("7/10", matrix["labels"])
        self.assertEqual(section["trend"]["chart"]["labels"], matrix["labels"])

    def test_new_pathfinder_without_targets_displays_placeholders(self) -> None:
        section = build_nev_section(
            "new-pathfinder",
            "2026款探陆",
            self.synthetic_report_date,
            self.synthetic_nev_daily["2026款探陆"],
            {},
            {},
        )
        cards = {card["label"]: card for card in section["summary"]["cards"]}
        matrix_rows = {row["key"]: row for row in section["trend"]["matrix"]["rows"]}

        self.assertEqual(cards["累计目标"]["displayValue"], "-")
        self.assertEqual(cards["累计达成率"]["displayValue"], "-")
        self.assertEqual(cards["当日目标"]["displayValue"], "-")
        self.assertEqual(cards["当日达成率"]["displayValue"], "-")
        self.assertTrue(all(value == "-" for value in matrix_rows["target"]["displayValues"]))
        self.assertTrue(all(value is None for value in section["trend"]["chart"]["series"]["target"]))

    def test_new_pathfinder_targets_are_used_when_available(self) -> None:
        report_date = date(2026, 7, 16)
        current_actuals = {
            date(2026, 7, 15): {"newLeads": 4, "arrivals": 1},
            report_date: {"newLeads": 6, "arrivals": 2},
        }
        current_targets = {
            date(2026, 7, 15): 5,
            report_date: 10,
        }

        section = build_nev_section(
            "new-pathfinder",
            "2026款探陆",
            report_date,
            current_actuals,
            {},
            current_targets,
        )
        cards = {card["label"]: card for card in section["summary"]["cards"]}

        self.assertEqual(cards["累计目标"]["displayValue"], "15")
        self.assertEqual(cards["累计达成率"]["displayValue"], "66.7%")
        self.assertEqual(cards["当日目标"]["displayValue"], "10")
        self.assertEqual(cards["当日达成率"]["displayValue"], "60.0%")

    def test_new_pathfinder_july_target_override_is_complete(self) -> None:
        report_date = date(2026, 7, 16)
        targets = resolve_new_pathfinder_targets(report_date, {})

        self.assertEqual(len(NEW_PATHFINDER_TARGET_OVERRIDES[(2026, 7)]), 31)
        self.assertEqual(sum(value or 0 for value in targets.values()), 7759)
        self.assertTrue(all(targets[date(2026, 7, day)] is None for day in range(1, 16)))
        self.assertEqual(sum((value or 0) for day, value in targets.items() if day <= report_date), 464)
        self.assertEqual(targets[report_date], 464)
        self.assertEqual(targets[date(2026, 7, 31)], 473)

        section = build_nev_section("new-pathfinder", "2026款探陆", report_date, {}, {}, targets)
        rows = {row["key"]: row for row in section["trend"]["matrix"]["rows"]}
        chart = section["trend"]["chart"]
        self.assertEqual(section["trend"]["matrix"]["labels"], [f"7/{day}" for day in range(16, 32)])
        self.assertTrue(all(len(row["displayValues"]) == 16 for row in rows.values()))
        self.assertEqual(rows["target"]["displayValues"][0], "464")
        self.assertEqual(rows["cumulativeTarget"]["displayValues"][0], "464")
        self.assertEqual(chart["labels"], [f"7/{day}" for day in range(16, 32)])
        self.assertEqual(chart["reportDayIndex"], 0)
        self.assertEqual(chart["series"]["target"][0], 464)
        self.assertEqual(chart["series"]["cumulativeTarget"][0], 464)

    def test_new_pathfinder_workbook_targets_take_precedence_over_override(self) -> None:
        report_date = date(2026, 7, 16)
        workbook_targets = {report_date: 999}

        self.assertIs(
            resolve_new_pathfinder_targets(report_date, workbook_targets),
            workbook_targets,
        )

    def test_nev_total_excludes_new_pathfinder_actuals(self) -> None:
        sections = {
            section["id"]: section
            for section in self.synthetic_payload["dashboards"]["nev"]["sections"]
        }
        sections_without_new_pathfinder = {
            section["id"]: section
            for section in self.synthetic_payload_without_new_pathfinder["dashboards"]["nev"][
                "sections"
            ]
        }
        new_pathfinder_cumulative = sections["new-pathfinder"]["summary"]["cards"][0]["value"]
        nev_total_cumulative = sections["nev-total"]["summary"]["cards"][0]["value"]
        nev_total_without_new_pathfinder = sections_without_new_pathfinder["nev-total"][
            "summary"
        ]["cards"][0]["value"]

        self.assertEqual(new_pathfinder_cumulative, 2)
        self.assertEqual(nev_total_cumulative, 10)
        self.assertEqual(nev_total_cumulative, nev_total_without_new_pathfinder)

    def test_lead_control_includes_new_pathfinder_valid_leads(self) -> None:
        lead_control = self.synthetic_payload["dashboards"]["lead-control"]["sections"][0]
        lead_control_without_new_pathfinder = self.synthetic_payload_without_new_pathfinder[
            "dashboards"
        ]["lead-control"]["sections"][0]
        report_index = self.synthetic_date.day - 1
        actual_all_vehicle_valid = lead_control["trend"]["chart"]["series"]["actual"][
            report_index
        ]
        actual_without_new_pathfinder = lead_control_without_new_pathfinder["trend"]["chart"][
            "series"
        ]["actual"][report_index]

        self.assertEqual(actual_all_vehicle_valid - actual_without_new_pathfinder, 1)

    def test_daily_brief_uses_separate_new_pathfinder_section(self) -> None:
        sections = self.synthetic_payload["dashboards"]["brief"]["briefing"]["sections"]
        sections_by_kind = {section["kind"]: section for section in sections}

        self.assertEqual(
            [section["kind"] for section in sections],
            ["intro", "valid-leads", "nev", "new-pathfinder", "arrival"],
        )
        self.assertEqual(sections_by_kind["valid-leads"]["title"], "全车系有效线索")
        self.assertEqual(sections_by_kind["nev"]["title"], "NEV新增线索")
        self.assertTrue(
            sections_by_kind["valid-leads"]["lines"][0].endswith(
                "（目标取值为H2穿透目标7月值）"
            )
        )
        self.assertNotIn("note", sections_by_kind["valid-leads"])
        self.assertEqual(
            sections_by_kind["nev"]["note"],
            "（目标取值为GTM输入的管控目标7月值）",
        )
        self.assertEqual(sections_by_kind["new-pathfinder"]["title"], "2026款探陆线索")
        self.assertIn("2026款探陆累计实绩 2", sections_by_kind["new-pathfinder"]["lines"][0])
        self.assertIn("累计达成率 -", sections_by_kind["new-pathfinder"]["lines"][0])
        self.assertFalse(any("2026款探陆" in line for line in sections_by_kind["nev"]["lines"]))

    def test_lead_control_row_order_is_stable(self) -> None:
        trend = self.payload["dashboards"]["lead-control"]["sections"][0]["trend"]
        self.assertEqual(
            trend["matrix"]["visibleRowKeys"],
            [
                "samePeriodActual",
                "samePeriodCumulative",
                "previousActual",
                "previousCumulative",
                "actual",
                "cumulativeActual",
                "dayYoy",
                "cumulativeYoy",
                "dayDelta",
                "cumulativeDelta",
            ],
        )

    def test_cumulative_actual_never_uses_na_placeholder(self) -> None:
        for dashboard in self.payload["dashboards"].values():
            for section in dashboard.get("sections", []):
                rows = section.get("trend", {}).get("matrix", {}).get("rows", [])
                for row in rows:
                    if row.get("key") == "cumulativeActual":
                        self.assertNotIn("#N/A", row.get("displayValues", []))

    def test_build_run_summary_contains_dashboard_counts(self) -> None:
        summary = build_run_summary(
            self.payload,
            LEADS_BOOK,
            ARRIVAL_BOOK,
            OUT_JSON,
            SUMMARY_JSON,
            False,
        )
        self.assertEqual(summary["reportDate"], self.payload["meta"]["reportDate"])
        self.assertEqual(summary["outputs"]["dashboardStatus"], "unchanged")
        self.assertEqual(summary["stats"]["dashboardCount"], 5)
        self.assertEqual(summary["stats"]["sectionCounts"]["lead-control"], 1)
        self.assertEqual(summary["inputs"]["workbookModifiedAt"], file_mtime_iso(LEADS_BOOK))
        self.assertEqual(summary["inputs"]["arrivalWorkbookModifiedAt"], file_mtime_iso(ARRIVAL_BOOK))

    def test_build_run_summary_includes_archive_outputs_when_provided(self) -> None:
        summary = build_run_summary(
            self.payload,
            LEADS_BOOK,
            ARRIVAL_BOOK,
            OUT_JSON,
            SUMMARY_JSON,
            True,
            archive_info={
                "monthKey": "2026-04",
                "dashboardPath": "./data/monthly/2026-04/dashboard.json",
                "summaryPath": "./data/monthly/2026-04/dashboard.summary.json",
                "indexPath": "./data/monthly/index.json",
                "dashboardChanged": True,
                "summaryChanged": False,
                "indexChanged": True,
            },
        )
        self.assertEqual(summary["outputs"]["archiveMonth"], "2026-04")
        self.assertEqual(summary["outputs"]["archiveIndexJson"], "./data/monthly/index.json")
        self.assertTrue(summary["outputs"]["archiveDashboardChanged"])

    def test_preserved_times_override_dashboard_and_summary_file_mtimes(self) -> None:
        preserved_times = {
            "workbookModifiedAt": "2026-07-14T09:06:37",
            "arrivalWorkbookModifiedAt": "2026-07-14T09:06:38",
        }
        payload = {
            **self.payload,
            "meta": {**self.payload["meta"]},
        }

        apply_preserved_input_modified_times(payload, preserved_times)
        summary = build_run_summary(
            payload,
            LEADS_BOOK,
            ARRIVAL_BOOK,
            OUT_JSON,
            SUMMARY_JSON,
            True,
            input_modified_times=preserved_times,
        )

        self.assertEqual(payload["meta"]["workbookModifiedAt"], preserved_times["workbookModifiedAt"])
        self.assertEqual(summary["inputs"]["workbookModifiedAt"], preserved_times["workbookModifiedAt"])
        self.assertEqual(
            summary["inputs"]["arrivalWorkbookModifiedAt"],
            preserved_times["arrivalWorkbookModifiedAt"],
        )

    def test_load_preserved_times_reads_consistent_committed_outputs(self) -> None:
        temp_dir = Path("tests/.tmp/preserved-input-times")
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            dashboard_path = temp_dir / "dashboard.json"
            summary_path = temp_dir / "dashboard.summary.json"
            dashboard_path.write_text(
                json.dumps({"meta": {"workbookModifiedAt": "2026-07-14T09:06:37"}}),
                encoding="utf-8",
            )
            summary_path.write_text(
                json.dumps(
                    {
                        "inputs": {
                            "workbookModifiedAt": "2026-07-14T09:06:37",
                            "arrivalWorkbookModifiedAt": "2026-07-14T09:06:38",
                        }
                    }
                ),
                encoding="utf-8",
            )

            preserved_times = load_preserved_input_modified_times(dashboard_path, summary_path)

            self.assertEqual(preserved_times["workbookModifiedAt"], "2026-07-14T09:06:37")
            self.assertEqual(preserved_times["arrivalWorkbookModifiedAt"], "2026-07-14T09:06:38")
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_load_preserved_times_rejects_inconsistent_leads_time(self) -> None:
        temp_dir = Path("tests/.tmp/inconsistent-preserved-input-times")
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            dashboard_path = temp_dir / "dashboard.json"
            summary_path = temp_dir / "dashboard.summary.json"
            dashboard_path.write_text(
                json.dumps({"meta": {"workbookModifiedAt": "2026-07-14T09:06:37"}}),
                encoding="utf-8",
            )
            summary_path.write_text(
                json.dumps(
                    {
                        "inputs": {
                            "workbookModifiedAt": "2026-07-14T01:07:00",
                            "arrivalWorkbookModifiedAt": "2026-07-14T09:06:38",
                        }
                    }
                ),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "dashboard 与 summary 中的线索工作簿时间不一致"):
                load_preserved_input_modified_times(dashboard_path, summary_path)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_load_preserved_times_rejects_missing_or_malformed_metadata(self) -> None:
        temp_dir = Path("tests/.tmp/malformed-preserved-input-times")
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            dashboard_path = temp_dir / "dashboard.json"
            summary_path = temp_dir / "dashboard.summary.json"
            dashboard_path.write_text(json.dumps({"meta": {}}), encoding="utf-8")
            summary_path.write_text(json.dumps({"inputs": {}}), encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "不是有效的 ISO 本地时间"):
                load_preserved_input_modified_times(dashboard_path, summary_path)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_pages_workflow_preserves_committed_input_modified_times(self) -> None:
        workflow = (Path(__file__).parents[1] / ".github/workflows/deploy-pages.yml").read_text(encoding="utf-8")
        self.assertIn("--preserve-input-modified-times", workflow)

    def test_apply_submission_time_normalizes_to_china_standard_time(self) -> None:
        payload = {"meta": {}}

        apply_submission_time(payload, "2026-09-03T07:43:28Z")

        self.assertEqual(payload["meta"]["submittedAt"], "2026-09-03T15:43:28")

    def test_arrival_dashboard_uses_nev_daily_arrivals_for_nev_actual_row(self) -> None:
        if self.previous_month_archive_payload is None:
            self.skipTest(f"missing monthly archive payload for {self.previous_month_key}")
        trend = self.previous_month_archive_payload["dashboards"]["arrival"]["sections"][0]["trend"]
        rows = {row["key"]: row["displayValues"] for row in trend["matrix"]["rows"]}
        report_index = trend["chart"]["reportDayIndex"]

        self.assertIn("nevActual", rows)
        self.assertNotEqual(rows["nevActual"][report_index], "-")

    def test_arrival_dashboard_keeps_first_day_for_ice_actual_row(self) -> None:
        if self.previous_month_archive_payload is None:
            self.skipTest(f"missing monthly archive payload for {self.previous_month_key}")
        trend = self.previous_month_archive_payload["dashboards"]["arrival"]["sections"][0]["trend"]
        rows = {row["key"]: row["displayValues"] for row in trend["matrix"]["rows"]}

        self.assertIn("iceActual", rows)
        self.assertNotEqual(rows["iceActual"][0], "-")

    def test_write_json_if_changed_ignores_generated_at_only(self) -> None:
        original = {
            "meta": {"generatedAt": "2026-04-15T16:00:00", "reportDate": "2026-04-15"},
            "dashboards": {"brief": {"id": "brief"}},
        }
        regenerated = {
            "meta": {"generatedAt": "2026-04-15T16:05:00", "reportDate": "2026-04-15"},
            "dashboards": {"brief": {"id": "brief"}},
        }

        temp_dir = Path("tests/.tmp/write-json-if-changed")
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            json_path = temp_dir / "dashboard.json"
            json_path.write_text(json.dumps(original, ensure_ascii=False, indent=2), encoding="utf-8")
            changed, existing_payload = write_json_if_changed(
                json_path,
                regenerated,
                volatile_field_paths=(("meta", "generatedAt"),),
            )

            self.assertFalse(changed)
            self.assertEqual(existing_payload, original)
            persisted = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(persisted["meta"]["generatedAt"], original["meta"]["generatedAt"])
        finally:
            shutil.rmtree(temp_dir.parent, ignore_errors=True)

    def test_write_monthly_archive_creates_snapshot_and_index(self) -> None:
        temp_dir = Path("tests/.tmp/monthly-archive")
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            archive_root = temp_dir / "monthly"
            index_path = archive_root / "index.json"
            expected_month = self.payload["meta"]["reportDate"][:7]
            summary_payload = build_run_summary(
                self.payload,
                LEADS_BOOK,
                ARRIVAL_BOOK,
                OUT_JSON,
                SUMMARY_JSON,
                True,
            )

            archive_info = write_monthly_archive(
                self.payload,
                summary_payload,
                archive_root=archive_root,
                index_path=index_path,
                docs_root=temp_dir,
            )

            self.assertEqual(archive_info["monthKey"], expected_month)
            self.assertTrue((archive_root / expected_month / "dashboard.json").exists())
            self.assertTrue((archive_root / expected_month / "dashboard.summary.json").exists())
            index_payload = json.loads(index_path.read_text(encoding="utf-8"))
            self.assertEqual(index_payload["latestMonth"], expected_month)
            self.assertEqual(index_payload["months"][0]["dashboardPath"], f"./monthly/{expected_month}/dashboard.json")
        finally:
            shutil.rmtree(temp_dir.parent, ignore_errors=True)


class BuildDashboardValidationTests(unittest.TestCase):
    def test_valid_leads_monthly_targets_load_from_versioned_config(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            config_path = Path(temporary_directory) / "dashboard_targets.json"
            config_path.write_text(
                json.dumps({"validLeadsMonthlyTargets": {"2026-07": 668262}}, ensure_ascii=False),
                encoding="utf-8",
            )

            actual = load_valid_leads_monthly_targets(config_path)

        self.assertEqual(actual, {(2026, 7): 668262})

    def test_valid_leads_monthly_targets_reject_invalid_values(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            config_path = Path(temporary_directory) / "dashboard_targets.json"
            config_path.write_text(
                json.dumps({"validLeadsMonthlyTargets": {"2026-07": 0}}, ensure_ascii=False),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "invalid valid leads target"):
                load_valid_leads_monthly_targets(config_path)

    def test_valid_leads_target_yoy_and_mom_use_their_expected_sources(self) -> None:
        report_date = date(2026, 7, 1)
        current = {report_date: {"validLeads": 334_131}}
        previous_period = {date(2026, 6, 1): {"validLeads": 320_000}}
        same_period = {date(2025, 7, 1): {"validLeads": 300_000}}

        trend = build_valid_leads_control_trend(
            report_date,
            current,
            previous_period,
            same_period,
            668_262,
        )
        items = {item["label"]: item["displayValue"] for item in trend["summary"]["items"]}
        brief = build_valid_leads_brief(
            report_date,
            current,
            previous_period,
            same_period,
            668_262,
        )

        self.assertEqual(items["累计实绩"], "334,131")
        self.assertEqual(items["本月目标"], "668,262")
        self.assertEqual(items["累计达成率"], "50.0%")
        self.assertEqual(items["同比"], "11.4%")
        self.assertEqual(items["环比"], "4.4%")
        self.assertEqual(
            brief["lines"],
            ["累计实绩 334,131，累计达成率 50.0%；同比 11.4%，环比 4.4%（目标取值为H2穿透目标7月值）"],
        )

    def test_valid_leads_brief_mom_matches_trend_for_month_to_date_period(self) -> None:
        report_date = date(2026, 8, 8)
        current = {date(2026, 8, 1): {"validLeads": 217_789}}
        previous_period = {
            date(2026, 7, 1): {"validLeads": 145_067},
            date(2026, 7, 31): {"validLeads": 548_862},
        }

        trend = build_valid_leads_control_trend(report_date, current, previous_period, {}, None)
        trend_items = {item["label"]: item["displayValue"] for item in trend["summary"]["items"]}
        brief = build_valid_leads_brief(report_date, current, previous_period, {}, None)

        self.assertEqual(trend_items["环比"], "50.1%")
        self.assertIn("环比 50.1%", brief["lines"][0])

    def test_valid_leads_full_same_month_keeps_month_to_date_yoy(self) -> None:
        report_date = date(2026, 9, 3)
        current = {date(2026, 9, day): {"validLeads": 20} for day in range(1, 4)}
        previous_period = {date(2026, 8, day): {"validLeads": 10} for day in range(1, 32)}
        same_period = {date(2025, 9, day): {"validLeads": 10} for day in range(1, 31)}

        trend = build_valid_leads_control_trend(report_date, current, previous_period, same_period, None)
        trend_items = {item["label"]: item["displayValue"] for item in trend["summary"]["items"]}
        brief = build_valid_leads_brief(report_date, current, previous_period, same_period, None)
        same_cumulative_row = next(
            row for row in trend["matrix"]["rows"] if row["key"] == "samePeriodCumulative"
        )

        self.assertEqual(trend_items["同比"], "100.0%")
        self.assertIn("同比 100.0%", brief["lines"][0])
        self.assertEqual(same_cumulative_row["displayValues"][2], "30")
        self.assertEqual(same_cumulative_row["displayValues"][-1], "300")

    def test_valid_leads_brief_mom_uses_previous_month_end_for_extra_day(self) -> None:
        report_date = date(2026, 7, 31)
        current = {date(2026, 7, 31): {"validLeads": 120}}
        previous_period = {
            date(2026, 6, 1): {"validLeads": 40},
            date(2026, 6, 30): {"validLeads": 60},
        }

        trend = build_valid_leads_control_trend(report_date, current, previous_period, {}, None)
        trend_items = {item["label"]: item["displayValue"] for item in trend["summary"]["items"]}
        brief = build_valid_leads_brief(report_date, current, previous_period, {}, None)

        self.assertEqual(trend_items["环比"], "20.0%")
        self.assertIn("环比 20.0%", brief["lines"][0])

    def test_arrival_brief_only_displays_previous_month_comparison(self) -> None:
        report_date = date(2026, 7, 2)
        arrival_maps = {
            "total_current": {date(2026, 7, 1): 100, report_date: 200},
            "total_same_period": {date(2025, 7, 1): 50, date(2025, 7, 2): 50},
            "total_previous_period": {date(2026, 6, 1): 80, date(2026, 6, 2): 120},
            "nev_current": {date(2026, 7, 1): 40, report_date: 60},
            "nev_same_period": {date(2025, 7, 1): 20, date(2025, 7, 2): 20},
            "nev_previous_period": {date(2026, 6, 1): 30, date(2026, 6, 2): 50},
            "ice_current": {date(2026, 7, 1): 60, report_date: 140},
            "ice_same_period": {date(2025, 7, 1): 30, date(2025, 7, 2): 30},
            "ice_previous_period": {date(2026, 6, 1): 50, date(2026, 6, 2): 70},
        }

        brief = build_arrival_brief(report_date, arrival_maps)

        self.assertEqual(
            brief["lines"][0],
            "全国累计来店 300，环比 50.0%；当日来店 200，环比 66.7%",
        )
        self.assertEqual(
            brief["lines"][1],
            "①NEV累计来店 100，环比 25.0%；当日来店 60，环比 20.0%",
        )
        self.assertEqual(
            brief["lines"][2],
            "②ICE累计来店 200，环比 66.7%；当日来店 140，环比 100.0%",
        )
        self.assertEqual(brief["sourceSheets"], ["NEV本期来店", "NEV上期来店", "ICE本期来店", "ICE上期来店"])

    def test_day_calendar_meta_distinguishes_holiday_weekend_makeup_and_regular_workday(self) -> None:
        self.assertEqual(get_day_calendar_meta(date(2026, 5, 4))["dayType"], "holiday")
        self.assertEqual(get_day_calendar_meta(date(2026, 5, 5))["dayType"], "holiday")
        self.assertEqual(get_day_calendar_meta(date(2026, 5, 9))["dayType"], "makeupWorkday")
        self.assertEqual(get_day_calendar_meta(date(2026, 5, 10))["dayType"], "weekend")
        self.assertEqual(get_day_calendar_meta(date(2026, 5, 11))["dayType"], "regularWorkday")

    def test_build_column_meta_marks_makeup_workday_without_weekend_or_holiday_flags(self) -> None:
        meta = build_column_meta(date(2026, 5, 9), date(2026, 4, 9))

        self.assertTrue(meta["highlightCurrent"])
        self.assertTrue(meta["isCurrentMakeupWorkday"])
        self.assertFalse(meta["isCurrentHoliday"])
        self.assertFalse(meta["isCurrentWeekend"])
        self.assertEqual(meta["currentDayType"], "makeupWorkday")

    def test_arrival_previous_cumulative_stops_at_last_available_previous_day(self) -> None:
        series = build_arrival_series(
            date(2026, 5, 5),
            {
                date(2026, 5, 1): 10,
                date(2026, 5, 2): 20,
                date(2026, 5, 3): 30,
                date(2026, 5, 4): 40,
                date(2026, 5, 5): 50,
            },
            {
                date(2025, 5, 1): 100,
                date(2025, 5, 2): 200,
                date(2025, 5, 3): 300,
            },
        )

        self.assertEqual(series["previousReportIndex"], 2)
        self.assertEqual(series["previousCumulative"][:6], [100, 300, 600, None, None, None])

    def test_arrival_full_previous_month_extends_chart_but_keeps_report_day_summary(self) -> None:
        report_date = date(2026, 9, 3)
        current = {date(2026, 9, day): 10 for day in range(1, 4)}
        previous_period = {date(2026, 8, day): day for day in range(1, 32)}
        same_period = {date(2025, 9, day): 5 for day in range(1, 31)}
        arrival_maps = {
            "total_current": current,
            "total_previous_period": previous_period,
            "total_same_period": same_period,
            "nev_current": current,
            "nev_previous_period": previous_period,
            "nev_same_period": same_period,
            "ice_current": {},
            "ice_previous_period": {},
            "ice_same_period": {},
        }

        dashboard = build_arrival_dashboard(report_date, arrival_maps)
        trend = dashboard["sections"][0]["trend"]
        summary = {item["label"]: item["value"] for item in trend["summary"]["items"]}

        self.assertEqual(trend["chart"]["series"]["previousActual"][:5], [1, 2, 3, 4, 5])
        self.assertEqual(trend["chart"]["series"]["previousActual"][-1], 30)
        self.assertEqual(trend["chart"]["series"]["previousCumulative"][-1], sum(range(1, 31)))
        self.assertEqual(summary["累计上期来店"], 6)
        self.assertEqual(summary["上期来店"], 3)

    def test_arrival_previous_cumulative_stays_empty_when_no_previous_data_exists(self) -> None:
        series = build_arrival_series(
            date(2026, 5, 1),
            {
                date(2026, 5, 1): 10,
            },
            {},
        )

        self.assertIsNone(series["previousReportIndex"])
        self.assertTrue(all(value is None for value in series["previousCumulative"]))

    def test_arrival_month_comparison_uses_previous_month_end_without_double_counting(self) -> None:
        series = build_arrival_series(
            date(2026, 7, 31),
            {date(2026, 7, day): 10 for day in range(1, 32)},
            {date(2026, 6, day): day for day in range(1, 31)},
            comparison_period="month",
        )

        self.assertEqual(series["previousDaily"][30], 30)
        self.assertEqual(series["previousCumulative"][29], 465)
        self.assertEqual(series["previousCumulative"][30], 465)
        self.assertEqual(series["previousReportIndex"], 30)

    def test_arrival_dashboard_populates_month_comparison_cards_on_extra_month_day(self) -> None:
        report_date = date(2026, 7, 31)
        current = {date(2026, 7, day): 10 for day in range(1, 32)}
        previous_period = {date(2026, 6, day): 5 for day in range(1, 31)}
        same_period = {date(2025, 7, day): 5 for day in range(1, 32)}
        arrival_maps = {
            "total_current": current,
            "total_previous_period": previous_period,
            "total_same_period": same_period,
            "nev_current": current,
            "nev_previous_period": previous_period,
            "nev_same_period": same_period,
            "ice_current": current,
            "ice_previous_period": previous_period,
            "ice_same_period": same_period,
        }

        dashboard = build_arrival_dashboard(report_date, arrival_maps)
        items = {
            item["label"]: item["displayValue"]
            for item in dashboard["sections"][0]["trend"]["summary"]["items"]
        }

        self.assertEqual(items["累计上期来店"], "150")
        self.assertEqual(items["累计环比"], "106.7%")
        self.assertEqual(items["上期来店"], "5")
        self.assertEqual(items["当日环比"], "100.0%")
        self.assertNotIn("累计同期来店", items)
        self.assertNotIn("累计同比", items)

        trend = dashboard["sections"][0]["trend"]
        self.assertEqual(trend["chartSubtitle"], "上期来店 / 本期来店 / 累计来店")
        self.assertEqual(trend["chart"]["seriesDefinitions"][0]["label"], "上期来店")
        self.assertEqual(trend["chart"]["seriesDefinitions"][-1]["label"], "累计上期来店")
        self.assertEqual(trend["chart"]["series"]["previousActual"][-1], 5)

    def test_load_arrival_daily_sheet_supports_sheets_without_header_row(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "2026年4月1日"
        sheet["B1"] = 1187
        sheet["A2"] = "2026年4月2日"
        sheet["B2"] = 1006

        actual = load_arrival_daily_sheet(sheet)

        self.assertEqual(list(actual.values())[:2], [1187, 1006])

    def test_validate_workbook_structure_requires_expected_sheets(self) -> None:
        leads = Workbook()
        arrival = Workbook()
        with self.assertRaisesRegex(ValueError, "线索工作簿 缺少必需工作表"):
            validate_workbook_structure(leads, arrival)

    def test_validate_sheet_headers_requires_expected_columns(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "目标竖版"
        sheet["A2"] = "不是合计"
        with self.assertRaisesRegex(ValueError, "缺少必需列"):
            validate_sheet_headers(sheet, 2, ("合计",), sheet.title)

    def test_validate_report_date_cell_uses_summary_fallback_when_formula_cache_missing(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "参数"
        sheet["C2"] = "not-a-date"
        self.assertIsNotNone(validate_report_date_cell(workbook))

    def test_validate_report_date_cell_requires_valid_date_without_fallback(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "参数"
        sheet["C2"] = "not-a-date"
        with patch("scripts.build_dashboard.read_json_file", return_value=None):
            with self.assertRaisesRegex(ValueError, "参数!C2 未读取到有效日期"):
                validate_report_date_cell(workbook)


if __name__ == "__main__":
    unittest.main()
