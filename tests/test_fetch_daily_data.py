from __future__ import annotations

import json
import shutil
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from scripts.build_dashboard import ARRIVAL_BOOK, LEADS_BOOK, safe_close_workbook
from scripts.fetch_daily_data import (
    ARRIVAL_SHEET_MAPPINGS,
    FETCH_TASKS,
    FetchTask,
    LEADS_SHEET_MAPPINGS,
    SHEET_MAPPINGS,
    parse_business_date,
    rebuild_dashboard,
    replace_workbook_sheets,
    resolve_export_path,
    run_fetch_task,
    run_update,
)


class FetchDailyDataTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_root = Path("tests/.tmp/fetch-daily-data")
        shutil.rmtree(self.temp_root, ignore_errors=True)
        self.temp_root.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_root, ignore_errors=True)

    def test_parse_business_date_supports_compact_format(self) -> None:
        self.assertEqual(parse_business_date("20260420"), date(2026, 4, 20))

    def test_daily_fetch_exports_same_period_leads_and_never_refills_sylphy_15(self) -> None:
        nev_task = next(task for task in FETCH_TASKS if task.label == "NEV 全国按日")
        ice_task = next(task for task in FETCH_TASKS if task.label == "ICE 全国按日")
        arrival_nev_task = next(task for task in FETCH_TASKS if task.label == "NEV 来店本期 + 上期 + 同期")

        self.assertEqual(nev_task.extra_args, ("--capture-wait-ms", "30000"))
        self.assertEqual(arrival_nev_task.extra_args, ("--safe-bootstrap", "--capture-wait-ms", "300000"))
        self.assertEqual(
            ice_task.report_keys,
            ("ice_national_daily", "ice_national_daily_same_period"),
        )
        self.assertEqual(
            [mapping.target_sheet for mapping in LEADS_SHEET_MAPPINGS],
            ["全国按日NEV", "全国按日NEV-同期", "全国按日ICE", "全国按日ICE-同期"],
        )
        self.assertNotIn(
            "十五代轩逸按日",
            [mapping.target_sheet for mapping in LEADS_SHEET_MAPPINGS],
        )
        period_suffix_sheets = [
            mapping.target_sheet for mapping in ARRIVAL_SHEET_MAPPINGS if mapping.allow_period_suffix
        ]
        self.assertEqual(
            period_suffix_sheets,
            ["NEV上期来店", "NEV同期来店", "ICE上期来店", "ICE同期来店"],
        )

    def test_resolve_export_path_uses_business_date_suffix(self) -> None:
        output_dir = self.temp_root / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "全国按日-0419.xlsx").write_text("old", encoding="utf-8")
        expected = output_dir / "全国按日-0420.xlsx"
        expected.write_text("new", encoding="utf-8")

        resolved = resolve_export_path(output_dir, "全国按日", date(2026, 4, 20))

        self.assertEqual(resolved, expected)

    def test_resolve_export_path_supports_alias_report_names(self) -> None:
        output_dir = self.temp_root / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)
        expected = output_dir / "专营店本期-0420.xlsx"
        expected.write_text("alias", encoding="utf-8")

        resolved = resolve_export_path(output_dir, ("NEV本期", "专营店本期"), date(2026, 4, 20))

        self.assertEqual(resolved, expected)

    def test_resolve_export_path_supports_nev_and_ice_comparison_period_suffixes(self) -> None:
        cases = (
            (("NEV上期", "专营店上期"), "NEV上期-0630.xlsx"),
            (("NEV同期", "专营店同期"), "NEV同期-0731.xlsx"),
            (("来店上期",), "来店上期-0630.xlsx"),
            (("来店同期",), "来店同期-0731.xlsx"),
        )
        for index, (report_names, filename) in enumerate(cases, start=1):
            with self.subTest(filename=filename):
                output_dir = self.temp_root / f"previous-period-{index}"
                output_dir.mkdir(parents=True, exist_ok=True)
                expected = output_dir / filename
                expected.write_text("previous", encoding="utf-8")

                with self.assertRaises(FileNotFoundError):
                    resolve_export_path(output_dir, report_names, date(2026, 7, 30))

                resolved = resolve_export_path(
                    output_dir,
                    report_names,
                    date(2026, 7, 30),
                    allow_period_suffix=True,
                )
                self.assertEqual(resolved, expected)

    def test_resolve_export_path_rejects_ambiguous_previous_period_files(self) -> None:
        output_dir = self.temp_root / "ambiguous-previous-period"
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "来店上期-0629.xlsx").write_text("older", encoding="utf-8")
        (output_dir / "来店上期-0630.xlsx").write_text("newer", encoding="utf-8")

        with self.assertRaisesRegex(RuntimeError, "导出文件匹配不唯一"):
            resolve_export_path(
                output_dir,
                "来店上期",
                date(2026, 7, 30),
                allow_period_suffix=True,
            )

    def test_run_fetch_task_retries_when_exporter_exits_zero_without_all_files(self) -> None:
        task = FetchTask(
            label="测试报表",
            script_path=self.temp_root / "fake_exporter.py",
            output_subdir="retry-check",
            report_keys=("first", "second"),
        )
        runtime_root = self.temp_root / "runtime"
        output_dir = runtime_root / task.output_subdir / "exports"
        attempts = 0

        def fake_stream(*_args, **_kwargs) -> None:
            nonlocal attempts
            attempts += 1
            if attempts == 2:
                (output_dir / "first-0420.xlsx").write_text("first", encoding="utf-8")
                (output_dir / "second-0420.xlsx").write_text("second", encoding="utf-8")

        with (
            patch("scripts.fetch_daily_data.stream_subprocess", side_effect=fake_stream),
            patch("scripts.fetch_daily_data.time.sleep"),
        ):
            resolved = run_fetch_task(
                task,
                business_date=date(2026, 4, 20),
                runtime_root=runtime_root,
                log=lambda _message: None,
                headless=True,
                username=None,
                password=None,
                chrome_path=None,
            )

        self.assertEqual(attempts, 2)
        self.assertEqual(resolved, output_dir)

    def test_run_fetch_task_does_not_retry_when_upstream_date_is_not_ready(self) -> None:
        task = FetchTask(
            label="NEV 来店",
            script_path=self.temp_root / "fake_exporter.py",
            output_subdir="upstream-date-check",
            report_keys=("current",),
        )
        error = RuntimeError("上游 NEV 来店数据尚未发布目标日期：2026-08-23")

        with (
            patch("scripts.fetch_daily_data.stream_subprocess", side_effect=error) as stream,
            patch("scripts.fetch_daily_data.time.sleep") as sleep,
            self.assertRaisesRegex(RuntimeError, "尚未发布目标日期"),
        ):
            run_fetch_task(
                task,
                business_date=date(2026, 8, 23),
                runtime_root=self.temp_root / "runtime",
                log=lambda _message: None,
                headless=True,
                username=None,
                password=None,
                chrome_path=None,
                max_attempts=3,
            )

        stream.assert_called_once()
        sleep.assert_not_called()

    def test_run_update_leads_only_fetches_and_replaces_four_leads_sheets(self) -> None:
        business_date = date(2026, 8, 24)
        runtime_dir = self.temp_root / "runtime"
        leads_path = self.temp_root / "NEV+ICE_xsai.xlsm"
        arrival_path = self.temp_root / "NEV+ICE_ldai.xlsm"
        fetched_tasks: list[str] = []

        def fake_run_fetch_task(task, **_kwargs):
            fetched_tasks.append(task.label)
            output_dir = runtime_dir / task.output_subdir / "exports"
            output_dir.mkdir(parents=True, exist_ok=True)
            export_names = {
                "NEV 全国按日": ("全国按日", "全国按日-同期"),
                "ICE 全国按日": ("全国按日ICE", "全国按日ICE-同期"),
            }[task.label]
            for export_name in export_names:
                (output_dir / f"{export_name}-0824.xlsx").write_text("export", encoding="utf-8")
            return output_dir

        with (
            patch("scripts.fetch_daily_data.build_runtime_dir", return_value=runtime_dir),
            patch("scripts.fetch_daily_data.run_fetch_task", side_effect=fake_run_fetch_task),
            patch("scripts.fetch_daily_data.replace_workbook_sheets") as replace_sheets,
            patch(
                "scripts.fetch_daily_data.rebuild_dashboard",
                return_value={"dashboardChanged": True, "summaryChanged": True},
            ) as rebuild,
        ):
            result = run_update(
                business_date=business_date,
                leads_path=leads_path,
                arrival_path=arrival_path,
                log=lambda _message: None,
                keep_runtime=True,
                leads_only=True,
            )

        self.assertEqual(fetched_tasks, ["NEV 全国按日", "ICE 全国按日"])
        replace_sheets.assert_called_once()
        replace_args = replace_sheets.call_args.args
        self.assertEqual(replace_args[0], leads_path)
        self.assertEqual(replace_args[2], LEADS_SHEET_MAPPINGS)
        self.assertEqual(set(replace_args[1]), {mapping.target_sheet for mapping in LEADS_SHEET_MAPPINGS})
        self.assertEqual(rebuild.call_args.kwargs["arrival_path"], arrival_path)
        self.assertEqual(result["updateScope"], "leads")
        self.assertEqual(set(result["exports"]), {mapping.result_label for mapping in LEADS_SHEET_MAPPINGS})

    def test_replace_workbook_sheets_overwrites_target_sheets(self) -> None:
        leads_path = self.temp_root / "NEV+ICE_xsai.xlsx"
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for mapping in SHEET_MAPPINGS:
            ws = workbook.create_sheet(mapping.target_sheet)
            ws["A1"] = "旧值"
            ws["B2"] = "待清空"
            ws.merge_cells("A1:B1")
        workbook.save(leads_path)
        workbook.close()

        export_paths: dict[str, Path] = {}
        for index, mapping in enumerate(SHEET_MAPPINGS, start=1):
            export_path = self.temp_root / f"{mapping.result_label}.xlsx"
            export_wb = Workbook()
            export_ws = export_wb.active
            export_ws.title = "导出"
            export_ws["A1"] = mapping.result_label
            export_ws["A2"] = index
            export_ws["C3"] = f"sheet-{index}"
            export_ws.merge_cells("A1:C1")
            export_wb.save(export_path)
            export_wb.close()
            export_paths[mapping.target_sheet] = export_path

        replace_workbook_sheets(leads_path, export_paths, SHEET_MAPPINGS, log=lambda _message: None)

        updated = load_workbook(leads_path)
        try:
            for index, mapping in enumerate(SHEET_MAPPINGS, start=1):
                ws = updated[mapping.target_sheet]
                self.assertEqual(ws["A1"].value, mapping.result_label)
                self.assertEqual(ws["A2"].value, index)
                self.assertEqual(ws["C3"].value, f"sheet-{index}")
                self.assertIn("A1:C1", {str(item) for item in ws.merged_cells.ranges})
                self.assertIsNone(ws["B2"].value)
        finally:
            safe_close_workbook(updated)

    def test_replace_workbook_sheets_supports_macro_and_non_macro_targets(self) -> None:
        macro_path = self.temp_root / "NEV+ICE_xsai.xlsm"
        arrival_path = self.temp_root / "NEV+ICE_ldai.xlsm"

        for workbook_path, mappings in ((macro_path, LEADS_SHEET_MAPPINGS), (arrival_path, ARRIVAL_SHEET_MAPPINGS)):
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            for mapping in mappings:
                ws = workbook.create_sheet(mapping.target_sheet)
                ws["A1"] = "旧值"
            workbook.save(workbook_path)
            workbook.close()

            export_paths: dict[str, Path] = {}
            for mapping in mappings:
                export_path = self.temp_root / f"{mapping.target_sheet}.xlsx"
                export_wb = Workbook()
                export_ws = export_wb.active
                export_ws.title = "导出"
                export_ws["A1"] = mapping.target_sheet
                export_wb.save(export_path)
                export_wb.close()
                export_paths[mapping.target_sheet] = export_path

            replace_workbook_sheets(
                workbook_path,
                export_paths,
                mappings,
                log=lambda _message: None,
                keep_vba=workbook_path.suffix.lower() == ".xlsm",
            )

            updated = load_workbook(workbook_path, keep_vba=workbook_path.suffix.lower() == ".xlsm")
            try:
                for mapping in mappings:
                    self.assertEqual(updated[mapping.target_sheet]["A1"].value, mapping.target_sheet)
            finally:
                safe_close_workbook(updated)

    def test_rebuild_dashboard_uses_business_date_override(self) -> None:
        out_path = self.temp_root / "dashboard.json"
        summary_path = self.temp_root / "dashboard.summary.json"

        result = rebuild_dashboard(
            business_date=date(2026, 4, 20),
            leads_path=LEADS_BOOK,
            arrival_path=ARRIVAL_BOOK,
            out_path=out_path,
            summary_path=summary_path,
            log=lambda _message: None,
            archive_root=self.temp_root / "monthly",
            archive_index_path=self.temp_root / "monthly" / "index.json",
            docs_root=self.temp_root,
        )

        payload = json.loads(out_path.read_text(encoding="utf-8"))
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        archive_payload = json.loads((self.temp_root / "monthly" / "2026-04" / "dashboard.json").read_text(encoding="utf-8"))
        archive_summary = json.loads((self.temp_root / "monthly" / "2026-04" / "dashboard.summary.json").read_text(encoding="utf-8"))
        archive_index = json.loads((self.temp_root / "monthly" / "index.json").read_text(encoding="utf-8"))
        self.assertEqual(payload["meta"]["reportDate"], "2026-04-20")
        self.assertEqual(summary["reportDate"], "2026-04-20")
        self.assertEqual(archive_payload["meta"]["reportDate"], "2026-04-20")
        self.assertEqual(archive_summary["reportDate"], "2026-04-20")
        self.assertEqual(archive_index["latestMonth"], "2026-04")
        self.assertIn("dashboardChanged", result)
        self.assertIn("summaryChanged", result)
        self.assertIn("archiveDashboardChanged", result)
        self.assertIn("archiveSummaryChanged", result)
        self.assertIn("archiveIndexChanged", result)


if __name__ == "__main__":
    unittest.main()
