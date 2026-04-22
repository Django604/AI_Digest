from __future__ import annotations

import copy
import importlib.util
import json
import re
import sys
import time
from datetime import date, datetime, timedelta
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent
DAILY_SOURCE_ROOT = WORKSPACE_ROOT / "日报取数平台"
ARRIVAL_NEV_DIR = DAILY_SOURCE_ROOT / "日报来店NEV源"
ARRIVAL_NEV_GETDATA = ARRIVAL_NEV_DIR / "getdata.py"
TARGET_REPORT_URL = "https://e3s-plus.dongfeng-nissan.com.cn/#/NEV-P-01/NEV-W-02/NEV-W-020104"
TARGET_CHART_WIDGET = "REPORT2"
TARGET_CHART_DATA_WIDGET = "自定义来店量"
TARGET_CHART_TAB_TEXT = "自定义"
TARGET_CHART_TAB_LAYOUT = "TABLAYOUT0"
TARGET_REPORT_KEYS = {
    "store_current_period",
    "store_previous_period",
    "store_same_period",
}
TARGET_REPORT_NAME = "来店批次分车系汇总表_按天"
QUERY_BUTTON_COORD = (1270, 252)
REPORT2_SEGMENT_WINDOW = 250_000
REPORT2_PATH_PATTERN = re.compile(r'class="vancharts-series-0 line".*?<path d="([^"]+)"', re.S)
REPORT2_POINT_PATTERN = re.compile(r"[ML]([0-9.]+),([0-9.]+)")
REPORT2_GRIDLINE_PATTERN = re.compile(r'<line y1="([0-9.]+)" y2="\1" x1="0" x2="[^"]+"')
REPORT2_AXIS_VALUE_PATTERN = re.compile(r'<text _x="([0-9.]+)" _y="([0-9.]+)"[^>]*>(-?\d+)</text>')
TARGET_PARAMETER_TEMPLATE = {
    "日期": "开始日期 ",
    "-": "结束日期 ",
    "专营店品牌-": "专营店品牌 ",
    "专营店品牌": "",
    "开始时间": "2026-04-01",
    "结束时间": "2026-04-21",
    "车型--": "车型 ",
    "车系": "",
    "专营店名称-": "专营店名称 ",
    "车辆品牌": "",
    "车辆品牌-": "车辆品牌 ",
    "客户来源-": "客户来源 ",
    "大区-": "大区 ",
    "大区": "",
    "小区-": "小区 ",
    "小区": "",
    "客户来源": "",
    "DLRCODE_TEXT": "",
    "DLRCODE": "",
    "LABEL0": "* 日期筛选仅对自定义报表生效",
    "渠道名称": "",
    "渠道名称-": "渠道名称 ",
}


def load_arrival_nev_module():
    if not ARRIVAL_NEV_GETDATA.exists():
        raise FileNotFoundError(f"未找到 NEV 来店取数脚本：{ARRIVAL_NEV_GETDATA}")

    for path in (ARRIVAL_NEV_DIR, DAILY_SOURCE_ROOT):
        text = str(path)
        if text not in sys.path:
            sys.path.insert(0, text)

    spec = importlib.util.spec_from_file_location("ai_digest_arrival_nev_getdata", ARRIVAL_NEV_GETDATA)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载模块：{ARRIVAL_NEV_GETDATA}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def patch_report_configs() -> None:
    report_configs_module = sys.modules.get("report_fetcher.report_configs")
    if report_configs_module is None:
        raise RuntimeError("未加载 report_fetcher.report_configs，无法修正 NEV 来店导出配置。")

    report_configs_module.REPORT_URL = TARGET_REPORT_URL
    report_configs = getattr(report_configs_module, "REPORT_CONFIGS", None)
    if not isinstance(report_configs, dict):
        raise RuntimeError("report_fetcher.report_configs.REPORT_CONFIGS 不可用，无法修正 NEV 来店导出配置。")

    for report_key in TARGET_REPORT_KEYS:
        config = report_configs.get(report_key)
        if not isinstance(config, dict):
            raise RuntimeError(f"NEV 来店导出配置缺少报表 key：{report_key}")
        config["report_url"] = TARGET_REPORT_URL
        config["report_tab"] = None
        config["parameterized_prepare_strategy"] = "interact_then_load"
        config["parameterized_prepare_parameters"] = copy.deepcopy(TARGET_PARAMETER_TEMPLATE)


def daterange(start_date: date, end_date: date) -> list[date]:
    current = start_date
    values: list[date] = []
    while current <= end_date:
        values.append(current)
        current += timedelta(days=1)
    return values


def extract_report2_segment(html_text: str) -> str:
    start_index = html_text.find(f'widgetname="{TARGET_CHART_WIDGET}"')
    if start_index < 0:
        raise RuntimeError("未在 iframe HTML 中找到 REPORT2，自定义来店图表尚未渲染。")

    end_index = html_text.find(f'widgetname="{TARGET_CHART_DATA_WIDGET}"', start_index)
    if end_index < 0:
        end_index = min(len(html_text), start_index + REPORT2_SEGMENT_WINDOW)
    return html_text[start_index:end_index]


def parse_report2_path_points(segment: str) -> list[tuple[float, float]]:
    match = REPORT2_PATH_PATTERN.search(segment)
    if match is None:
        raise RuntimeError("未在 REPORT2 片段中找到折线路径，无法提取自定义来店数据。")
    return [(float(x), float(y)) for x, y in REPORT2_POINT_PATTERN.findall(match.group(1))]


def resolve_report2_axis_scale(segment: str) -> tuple[float, float, int, int]:
    gridlines = [float(value) for value in REPORT2_GRIDLINE_PATTERN.findall(segment)]
    if not gridlines:
        raise RuntimeError("未在 REPORT2 片段中识别到图表网格线，无法反推数值刻度。")

    axis_values = [
        int(value_text)
        for x_text, _y_text, value_text in REPORT2_AXIS_VALUE_PATTERN.findall(segment)
        if float(x_text) < 50
    ]
    if not axis_values:
        raise RuntimeError("未在 REPORT2 片段中识别到图表纵轴刻度，无法反推数值刻度。")

    return min(gridlines), max(gridlines), min(axis_values), max(axis_values)


def parse_report2_daily_series(html_text: str, start_date: date, end_date: date) -> list[tuple[date, int]]:
    segment = extract_report2_segment(html_text)
    points = parse_report2_path_points(segment)
    dates = daterange(start_date, end_date)
    if len(points) != len(dates):
        raise RuntimeError(
            f"REPORT2 点位数与目标日期数不一致：points={len(points)} dates={len(dates)}。"
        )

    plot_top, plot_bottom, axis_min_value, axis_max_value = resolve_report2_axis_scale(segment)
    if plot_bottom <= plot_top or axis_max_value == axis_min_value:
        raise RuntimeError("REPORT2 图表刻度异常，无法换算自定义来店数据。")

    scale_span = plot_bottom - plot_top
    value_span = axis_max_value - axis_min_value
    result: list[tuple[date, int]] = []
    for current_date, (_x_value, y_value) in zip(dates, points):
        raw_value = axis_min_value + ((plot_bottom - y_value) / scale_span) * value_span
        result.append((current_date, int(round(raw_value))))
    return result


def write_daily_excel(output_dir: Path, report_name: str, daily_rows: list[tuple[date, int]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "report0"
    worksheet["A1"] = "日期"
    worksheet["B1"] = "来店量"
    for row_index, (current_date, value) in enumerate(daily_rows, start=2):
        worksheet.cell(row=row_index, column=1, value=current_date)
        worksheet.cell(row=row_index, column=2, value=value)

    excel_path = output_dir / f"{report_name}-{build_business_suffix(daily_rows[-1][0])}.xlsx"
    workbook.save(excel_path)
    workbook.close()
    return excel_path


def build_business_suffix(value: date) -> str:
    return value.strftime("%m%d")


def build_parameterized_interact_post_data(parameters: dict) -> str:
    parameter_text = json.dumps(parameters, ensure_ascii=False, separators=(",", ":"))
    return f"__parameters__={parameter_text}&__widgetname__=[]"


def build_parameterized_submit_post_data(parameters: dict) -> str:
    parameter_payload = dict(parameters)
    for key in ("QUERY", "PAGINGQUERY", "隐藏"):
        if key in parameter_payload:
            parameter_payload[key] = "0"
    parameter_text = json.dumps(parameter_payload, ensure_ascii=False, separators=(",", ":"))
    return f"__parameters__={parameter_text}"


def build_widget_load_post_data(widget_name: str, export_page_size: int) -> str:
    return (
        f"widgetName={widget_name}"
        "&pageIndex=1"
        f"&__parameters__={{query:'1',pageIndex:'1',pageSize:'{export_page_size}'}}"
        "&noCache=true"
        "&simpleJson=false"
        "&arrayJson=false"
    )


def build_tab_execute_post_data(tab_name: str, tab_layout_name: str = TARGET_CHART_TAB_LAYOUT) -> str:
    return f"tabName={tab_name}&tabLayoutName={tab_layout_name}"


def rewrite_timestamp_query(url: str) -> str:
    parsed = urlsplit(url)
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    rewritten = []
    now_value = str(int(time.time() * 1000))
    changed = False
    for key, value in pairs:
        if key == "_":
            rewritten.append((key, now_value))
            changed = True
        else:
            rewritten.append((key, value))
    if not changed:
        rewritten.append(("_", now_value))
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urlencode(rewritten), parsed.fragment))


def replace_url_path(url: str, target_suffix: str) -> str:
    parsed = urlsplit(url)
    path = parsed.path
    marker = "/view/fit/form/"
    marker_index = path.find(marker)
    if marker_index < 0:
        raise RuntimeError(f"无法从 URL 推断 FineReport 接口路径：{url}")
    base_path = path[: marker_index + len(marker)]
    return rewrite_timestamp_query(
        urlunsplit(
            (parsed.scheme, parsed.netloc, f"{base_path}{target_suffix}", parsed.query, parsed.fragment)
        )
    )


def parse_widget_prepare_payload(raw_text: str) -> list[list[dict]]:
    payload = json.loads(raw_text)
    if isinstance(payload, dict) and isinstance(payload.get("pageResult"), list):
        return payload["pageResult"]
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict) and isinstance(item.get("pageResult"), list):
                return item["pageResult"]
    raise RuntimeError("load/content 响应中未找到 pageResult，无法提取自定义来店数据。")


def decode_fr_cell_text(cell: dict) -> str:
    raw_value = cell.get("value")
    if raw_value in (None, ""):
        return ""
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if text.startswith("{") and text.endswith("}"):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                return text
            if isinstance(parsed, dict):
                return str(parsed.get("value", "")).strip()
        return text
    return str(raw_value).strip()


def parse_date_text(value: str) -> date | None:
    text = value.strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return date.fromisoformat(text) if fmt == "%Y-%m-%d" else datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_int_text(value: str) -> int | None:
    text = value.strip().replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


@dataclass(frozen=True)
class SimpleChartMeta:
    chart_id: str
    ec_name: str


def parse_fr_cell_value(cell: dict):
    raw_value = cell.get("value")
    if not isinstance(raw_value, str):
        return raw_value
    text = raw_value.strip()
    if not text:
        return text
    if text.startswith("{") and text.endswith("}"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def normalize_chart_ec_name(value: str) -> str:
    text = value.strip()
    if not text:
        return TARGET_CHART_WIDGET
    if text.lower().startswith("report"):
        suffix = text[len("report") :]
        return f"REPORT{suffix}"
    return text


def extract_simplechart_meta_from_page_result(page_result: list[list[dict]]) -> SimpleChartMeta:
    for row in page_result:
        for cell in row:
            parsed_value = parse_fr_cell_value(cell)
            if not isinstance(parsed_value, dict) or parsed_value.get("type") != "simplechart":
                continue
            items = parsed_value.get("items")
            if not isinstance(items, list):
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                chart_id = str(item.get("simpleChartInShowID") or "").strip()
                url_text = str(item.get("url") or "").strip()
                if not chart_id and not url_text:
                    continue
                query_pairs = dict(parse_qsl(urlsplit(url_text).query, keep_blank_values=True))
                chart_id = chart_id or str(query_pairs.get("chartID") or "").strip()
                ec_name = normalize_chart_ec_name(str(query_pairs.get("ecName") or TARGET_CHART_WIDGET))
                if chart_id:
                    return SimpleChartMeta(chart_id=chart_id, ec_name=ec_name)
    raise RuntimeError("未在 REPORT2 的 pageResult 中找到 simplechart，无法继续请求 chart.data。")


def build_chart_data_url(prepare_url: str, session_id: str, chart_id: str, ec_name: str) -> str:
    parsed = urlsplit(replace_url_path(prepare_url, "chart/data"))
    query_pairs = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query_pairs.update(
        {
            "sessionID": session_id,
            "chartID": chart_id,
            "ecName": ec_name,
        }
    )
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urlencode(query_pairs), parsed.fragment))


def extract_daily_rows_from_chart_payload(chart_payload: dict, start_date: date, end_date: date) -> list[tuple[date, int]]:
    series_list = chart_payload.get("chartAttr", {}).get("series")
    if not isinstance(series_list, list) or not series_list:
        raise RuntimeError("chart.data 响应缺少 chartAttr.series，无法提取自定义来店按日数据。")

    values_by_date: dict[date, int] = {}
    for series in series_list:
        if not isinstance(series, dict):
            continue
        data_items = series.get("data")
        if not isinstance(data_items, list):
            continue
        for item in data_items:
            if not isinstance(item, dict):
                continue
            current_date = parse_date_text(str(item.get("originalCategory") or item.get("x") or ""))
            current_value = parse_int_text(str(item.get("y") or item.get("value") or ""))
            if current_date is None or current_value is None:
                continue
            if start_date <= current_date <= end_date:
                values_by_date[current_date] = current_value

    expected_dates = daterange(start_date, end_date)
    missing_dates = [item.isoformat() for item in expected_dates if item not in values_by_date]
    if missing_dates:
        raise RuntimeError(
            "chart.data 返回的自定义来店按日数据缺少日期："
            + ", ".join(missing_dates[:5])
            + (" ..." if len(missing_dates) > 5 else "")
        )
    return [(current_date, values_by_date[current_date]) for current_date in expected_dates]


def extract_daily_rows_from_page_result(page_result: list[list[dict]], start_date: date, end_date: date) -> list[tuple[date, int]]:
    values_by_date: dict[date, int] = {}
    for row in page_result:
        ordered_cells = sorted(row, key=lambda cell: int(cell.get("position", {}).get("x", 0)))
        row_values = [decode_fr_cell_text(cell) for cell in ordered_cells]
        if len(row_values) < 2:
            continue
        current_date = parse_date_text(row_values[0])
        current_value = parse_int_text(row_values[1])
        if current_date is None or current_value is None:
            continue
        if start_date <= current_date <= end_date:
            values_by_date[current_date] = current_value

    expected_dates = daterange(start_date, end_date)
    missing_dates = [item.isoformat() for item in expected_dates if item not in values_by_date]
    if missing_dates:
        raise RuntimeError(
            "自定义来店按日数据缺少日期："
            + ", ".join(missing_dates[:5])
            + (" ..." if len(missing_dates) > 5 else "")
        )
    return [(current_date, values_by_date[current_date]) for current_date in expected_dates]


def bootstrap_export_context(module, args, context, page, output_dir: Path):
    bootstrap_runner = getattr(module, "_bootstrap_group_runtime", None)
    trace_dir = output_dir / "_trace" / "_page_bootstrap_1"
    if callable(bootstrap_runner):
        return bootstrap_runner(
            context=context,
            page=page,
            report_url=TARGET_REPORT_URL,
            trace_dir=trace_dir,
            args=args,
        )
    return module.bootstrap_finereport_export_context(
        context=context,
        page=page,
        report_url=TARGET_REPORT_URL,
        capture_wait_ms=int(args.capture_wait_ms),
        trace_dir=trace_dir,
        save_body=args.save_body,
        fast_bootstrap=bool(args.fast_bootstrap),
        recorder_cls=module.ResponseRecorder,
        persist_capture_artifacts=module.persist_capture_artifacts,
        build_parameterized_export_runtime_context=module.build_parameterized_export_runtime_context,
        open_new_retail_system_for_capture=module.open_new_retail_system_for_capture,
        wait_for_parameterized_export_context=module.wait_for_parameterized_export_context,
        log=module.log,
    )


def capture_custom_chart_series_via_api(export_context, filter_config, trace_dir: Path) -> list[tuple[date, int]]:
    export_page_size = filter_config.export_page_size or 150000
    parameters = filter_config.parameterized_prepare_parameters or {}
    submit_data = build_parameterized_submit_post_data(parameters)
    interact_data = build_parameterized_interact_post_data(parameters)
    load_data = build_widget_load_post_data(TARGET_CHART_WIDGET, export_page_size)
    tab_execute_data = build_tab_execute_post_data(TARGET_CHART_TAB_TEXT)
    tab_execute_url = replace_url_path(export_context.prepare_url, "tab/execute")
    start_date = date.fromisoformat(filter_config.start_date)
    end_date = date.fromisoformat(filter_config.end_date)

    submit_response = export_context.session.post(
        rewrite_timestamp_query(export_context.submit_url),
        headers=export_context.prepare_headers,
        data=submit_data,
        timeout=300,
    )
    submit_response.raise_for_status()

    interact_response = export_context.session.post(
        rewrite_timestamp_query(export_context.interact_url),
        headers=export_context.prepare_headers,
        data=interact_data,
        timeout=300,
    )
    interact_response.raise_for_status()

    tab_execute_response = export_context.session.post(
        tab_execute_url,
        headers=export_context.prepare_headers,
        data=tab_execute_data,
        timeout=300,
    )
    tab_execute_response.raise_for_status()

    load_response = export_context.session.post(
        rewrite_timestamp_query(export_context.prepare_url),
        headers=export_context.prepare_headers,
        data=load_data,
        timeout=300,
    )
    load_response.raise_for_status()

    trace_dir.mkdir(parents=True, exist_ok=True)
    load_response_path = trace_dir / "report2_load_response.json"
    load_response_path.write_text(load_response.text, encoding="utf-8")
    (trace_dir / "report2_api_trace.json").write_text(
        json.dumps(
            {
                "widget_name": TARGET_CHART_WIDGET,
                "submit_url": export_context.submit_url,
                "submit_status": submit_response.status_code,
                "interact_url": export_context.interact_url,
                "interact_status": interact_response.status_code,
                "tab_execute_url": tab_execute_url,
                "tab_execute_status": tab_execute_response.status_code,
                "tab_execute_post_data": tab_execute_data,
                "prepare_url": export_context.prepare_url,
                "prepare_status": load_response.status_code,
                "prepare_post_data": load_data,
                "prepare_response_path": str(load_response_path),
                "response_preview": load_response.text[:4000],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    page_result = parse_widget_prepare_payload(load_response.text)
    try:
        return extract_daily_rows_from_page_result(page_result, start_date, end_date)
    except RuntimeError as page_result_error:
        chart_meta = extract_simplechart_meta_from_page_result(page_result)
        chart_data_url = build_chart_data_url(
            export_context.prepare_url,
            export_context.session_id,
            chart_meta.chart_id,
            chart_meta.ec_name,
        )
        chart_response = export_context.session.get(
            chart_data_url,
            headers=export_context.export_headers,
            timeout=300,
        )
        chart_response.raise_for_status()
        chart_response_path = trace_dir / "report2_chart_data.json"
        chart_response_path.write_text(chart_response.text, encoding="utf-8")
        (trace_dir / "report2_chart_trace.json").write_text(
            json.dumps(
                {
                    "chart_id": chart_meta.chart_id,
                    "ec_name": chart_meta.ec_name,
                    "chart_data_url": chart_data_url,
                    "chart_status": chart_response.status_code,
                    "chart_response_path": str(chart_response_path),
                    "fallback_reason": str(page_result_error),
                    "chart_response_preview": chart_response.text[:4000],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        chart_payload = json.loads(chart_response.text)
        if not isinstance(chart_payload, dict):
            raise RuntimeError("chart.data 响应不是 JSON 对象，无法提取自定义来店按日数据。")
        return extract_daily_rows_from_chart_payload(chart_payload, start_date, end_date)


def click_query_button(page, frame) -> None:
    candidates = (
        frame.get_by_role("button", name="查询"),
        frame.locator("button:has-text('查询')"),
        frame.locator("[role='button']:has-text('查询')"),
        frame.get_by_text("查询", exact=False),
        page.get_by_role("button", name="查询"),
        page.locator("button:has-text('查询')"),
        page.locator("[role='button']:has-text('查询')"),
        page.get_by_text("查询", exact=False),
    )
    for locator in candidates:
        try:
            target = locator.first
            if target.is_visible(timeout=1_500):
                target.click(timeout=3_000)
                return
        except Exception:
            continue
    page.mouse.click(*QUERY_BUTTON_COORD)


def click_custom_chart_tab(page, frame) -> bool:
    candidates = (
        frame.get_by_text(TARGET_CHART_TAB_TEXT, exact=True),
        frame.locator(f"button:has-text('{TARGET_CHART_TAB_TEXT}')"),
        frame.locator(f"[role='button']:has-text('{TARGET_CHART_TAB_TEXT}')"),
        page.get_by_text(TARGET_CHART_TAB_TEXT, exact=True),
        page.locator(f"button:has-text('{TARGET_CHART_TAB_TEXT}')"),
        page.locator(f"[role='button']:has-text('{TARGET_CHART_TAB_TEXT}')"),
    )
    for locator in candidates:
        try:
            target = locator.first
            if target.is_visible(timeout=1_500):
                target.click(timeout=3_000, force=True)
                return True
        except Exception:
            continue

    try:
        clicked = frame.evaluate(
            """
            () => {
              const exactNode = document.evaluate(
                "//*[normalize-space(text())='自定义']",
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
              ).singleNodeValue;
              if (exactNode && typeof exactNode.click === 'function') {
                exactNode.click();
                return true;
              }
              const fuzzyNode = Array.from(document.querySelectorAll('*')).find((node) => {
                const text = (node.textContent || '').trim();
                return text === '自定义';
              });
              if (fuzzyNode && typeof fuzzyNode.click === 'function') {
                fuzzyNode.click();
                return true;
              }
              return false;
            }
            """
        )
        return bool(clicked)
    except Exception:
        return False


def is_custom_chart_html_ready(html_text: str, expected_start: str, expected_end: str) -> bool:
    return (
        f'widgetname="{TARGET_CHART_WIDGET}"' in html_text
        and expected_start in html_text
        and expected_end in html_text
        and 'vancharts-series-0 line' in html_text
    )


def wait_for_custom_chart_html(page, frame, start_date: date, end_date: date, *, timeout_ms: int = 120_000) -> str:
    expected_start = start_date.strftime("%Y-%m-%d")
    expected_end = end_date.strftime("%Y-%m-%d")
    waited_ms = 0
    while waited_ms <= timeout_ms:
        html_text = frame.content()
        if is_custom_chart_html_ready(html_text, expected_start, expected_end):
            return html_text
        if waited_ms in (0, 4_000, 12_000, 24_000):
            click_custom_chart_tab(page, frame)
        frame.wait_for_timeout(2_000)
        waited_ms += 2_000
    raise RuntimeError(
        f"等待自定义来店图表超时：未在 {timeout_ms / 1000:.0f} 秒内等到 {expected_start} ~ {expected_end} 的 REPORT2。"
    )


def capture_custom_chart_series(module, datetest_module, page, filter_config) -> list[tuple[date, int]]:
    page.goto(TARGET_REPORT_URL, wait_until="domcontentloaded", timeout=60_000)
    module.wait_for_page_ready(page)
    page.wait_for_timeout(3_000)
    datetest_module.set_date_range(page, filter_config.start_date, filter_config.end_date)
    page.wait_for_timeout(1_500)
    frame = datetest_module.get_report_frame(page)
    click_custom_chart_tab(page, frame)
    page.wait_for_timeout(1_000)
    click_query_button(page, frame)
    page.wait_for_timeout(1_000)
    click_custom_chart_tab(page, frame)
    html_text = wait_for_custom_chart_html(
        page,
        frame,
        date.fromisoformat(filter_config.start_date),
        date.fromisoformat(filter_config.end_date),
    )
    return parse_report2_daily_series(
        html_text,
        date.fromisoformat(filter_config.start_date),
        date.fromisoformat(filter_config.end_date),
    )


def main() -> int:
    module = load_arrival_nev_module()
    patch_report_configs()

    args = module.build_parser().parse_args()
    module.apply_business_date_override(args.business_date, module.normalize_business_date_text)
    chrome_path = module.find_local_chrome_executable(args.chrome_path)
    output_dir = module.build_output_dir(Path(args.output_dir), args.output_folder_name, args.business_date)
    output_dir.mkdir(parents=True, exist_ok=True)
    report_configs = module.build_report_configs(args)
    if not report_configs:
        raise RuntimeError("当前没有可执行的 NEV 来店报表配置。")

    with module.sync_playwright() as playwright:
        browser, context, page = module.create_browser_context(
            playwright,
            chrome_path,
            headless=args.headless,
            slow_mo=args.slow_mo,
        )
        module.ensure_logged_in_portal(page, args)
        bootstrap_result = bootstrap_export_context(module, args, context, page, output_dir)
        export_context = bootstrap_result.export_context

        success_count = 0
        try:
            for index, filter_config in enumerate(report_configs, start=1):
                module.log(f"[{index:04d} {filter_config.report_name}] 开始通过后台接口抓取 REPORT2 自定义来店按日数据。")
                daily_rows = capture_custom_chart_series_via_api(
                    export_context,
                    filter_config,
                    output_dir / "_trace" / filter_config.key,
                )
                excel_path = write_daily_excel(output_dir, filter_config.report_name, daily_rows)
                module.log(f"[{index:04d} {filter_config.report_name}] Excel 已保存到：{excel_path}")
                success_count += 1
        finally:
            try:
                context.close()
            finally:
                browser.close()

    return 0 if success_count == len(report_configs) else 1


if __name__ == "__main__":
    raise SystemExit(main())
