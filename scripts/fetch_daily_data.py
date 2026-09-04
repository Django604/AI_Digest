from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable

from copy import copy
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_dashboard import (  # noqa: E402
    ARRIVAL_BOOK,
    DOCS_DIR,
    LEADS_BOOK,
    MONTHLY_ARCHIVE_DIR,
    MONTHLY_ARCHIVE_INDEX,
    OUT_JSON,
    SUMMARY_JSON,
    build_payload,
    build_run_summary,
    coerce_date,
    safe_close_workbook,
    write_json_if_changed,
    write_monthly_archive,
)


DAILY_SOURCE_ROOT = WORKSPACE_ROOT / "日报取数平台"
NEV_SCRIPT = DAILY_SOURCE_ROOT / "日报线索NEV源" / "getdata.py"
LEADS_NEV_WRAPPER_SCRIPT = PROJECT_ROOT / "scripts" / "run_leads_nev_exports.py"
LEADS_ICE_WRAPPER_SCRIPT = PROJECT_ROOT / "scripts" / "run_leads_ice_exports.py"
ARRIVAL_ICE_SCRIPT = PROJECT_ROOT / "scripts" / "run_arrival_ice_exports.py"
ARRIVAL_NEV_WRAPPER_SCRIPT = PROJECT_ROOT / "scripts" / "run_arrival_nev_exports.py"
RUNTIME_ROOT = PROJECT_ROOT / ".runtime" / "daily_update"
UPSTREAM_DATE_NOT_READY_MARKER = "上游 NEV 来店数据尚未发布目标日期"


@dataclass(frozen=True)
class FetchTask:
    label: str
    script_path: Path
    output_subdir: str
    report_keys: tuple[str, ...]
    extra_args: tuple[str, ...] = ()


@dataclass(frozen=True)
class SheetUpdateMapping:
    export_names: tuple[str, ...]
    result_label: str
    target_sheet: str
    workbook_kind: str
    allow_period_suffix: bool = False


LEADS_WORKBOOK_KIND = "leads"
ARRIVAL_WORKBOOK_KIND = "arrival"


LEADS_FETCH_TASKS = (
    FetchTask(
        label="NEV 全国按日",
        script_path=LEADS_NEV_WRAPPER_SCRIPT,
        output_subdir="nev",
        report_keys=("national_daily", "national_daily_same_period"),
        extra_args=("--capture-wait-ms", "30000"),
    ),
    FetchTask(
        label="ICE 全国按日",
        script_path=LEADS_ICE_WRAPPER_SCRIPT,
        output_subdir="ice",
        report_keys=("ice_national_daily", "ice_national_daily_same_period"),
    ),
)

ARRIVAL_FETCH_TASKS = (
    FetchTask(
        label="NEV 来店本期 + 上期 + 同期",
        script_path=ARRIVAL_NEV_WRAPPER_SCRIPT,
        output_subdir="arrival-nev",
        report_keys=("store_current_period", "store_previous_period", "store_same_period"),
        extra_args=("--safe-bootstrap", "--capture-wait-ms", "300000"),
    ),
    FetchTask(
        label="ICE 来店本期 + 上期 + 同期",
        script_path=ARRIVAL_ICE_SCRIPT,
        output_subdir="arrival-ice",
        report_keys=(
            "store_batch_vehicle_summary_本期_来店",
            "store_batch_vehicle_summary_上期_来店",
            "store_batch_vehicle_summary_同期_来店",
        ),
    ),
)

FETCH_TASKS = LEADS_FETCH_TASKS + ARRIVAL_FETCH_TASKS

LEADS_SHEET_MAPPINGS = (
    SheetUpdateMapping(export_names=("全国按日",), result_label="全国按日", target_sheet="全国按日NEV", workbook_kind=LEADS_WORKBOOK_KIND),
    SheetUpdateMapping(export_names=("全国按日-同期",), result_label="全国按日-同期", target_sheet="全国按日NEV-同期", workbook_kind=LEADS_WORKBOOK_KIND),
    SheetUpdateMapping(export_names=("全国按日ICE",), result_label="全国按日ICE", target_sheet="全国按日ICE", workbook_kind=LEADS_WORKBOOK_KIND),
    SheetUpdateMapping(export_names=("全国按日ICE-同期",), result_label="全国按日ICE-同期", target_sheet="全国按日ICE-同期", workbook_kind=LEADS_WORKBOOK_KIND),
)

ARRIVAL_SHEET_MAPPINGS = (
    SheetUpdateMapping(
        export_names=("NEV本期", "专营店本期"),
        result_label="NEV本期来店",
        target_sheet="NEV本期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
    ),
    SheetUpdateMapping(
        export_names=("NEV上期", "专营店上期"),
        result_label="NEV上期来店",
        target_sheet="NEV上期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
        allow_period_suffix=True,
    ),
    SheetUpdateMapping(
        export_names=("NEV同期", "专营店同期"),
        result_label="NEV同期来店",
        target_sheet="NEV同期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
        allow_period_suffix=True,
    ),
    SheetUpdateMapping(
        export_names=("来店本期",),
        result_label="ICE本期来店",
        target_sheet="ICE本期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
    ),
    SheetUpdateMapping(
        export_names=("来店上期",),
        result_label="ICE上期来店",
        target_sheet="ICE上期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
        allow_period_suffix=True,
    ),
    SheetUpdateMapping(
        export_names=("来店同期",),
        result_label="ICE同期来店",
        target_sheet="ICE同期来店",
        workbook_kind=ARRIVAL_WORKBOOK_KIND,
        allow_period_suffix=True,
    ),
)

SHEET_MAPPINGS = (*LEADS_SHEET_MAPPINGS, *ARRIVAL_SHEET_MAPPINGS)


def default_log(message: str) -> None:
    print(message, flush=True)


def parse_business_date(value: str | None = None) -> date:
    if value:
        parsed = coerce_date(value)
        if parsed is None:
            raise ValueError(f"无法识别业务日期：{value}")
        return parsed
    return date.today() - timedelta(days=1)


def format_business_date(value: date) -> str:
    return value.strftime("%Y-%m-%d")


def build_business_suffix(value: date) -> str:
    return value.strftime("%m%d")


def build_task_output_dir(task: FetchTask, run_root: Path) -> Path:
    return run_root / task.output_subdir / "exports"


def resolve_export_path(
    output_dir: Path,
    report_name: str | tuple[str, ...],
    business_date: date,
    *,
    allow_period_suffix: bool = False,
) -> Path:
    suffix = build_business_suffix(business_date)
    report_names = (report_name,) if isinstance(report_name, str) else report_name
    for candidate_name in report_names:
        candidates = sorted(output_dir.glob(f"{candidate_name}-{suffix}.*"))
        if candidates:
            return candidates[0]

    if allow_period_suffix:
        period_candidates: set[Path] = set()
        for candidate_name in report_names:
            name_prefix = f"{candidate_name}-"
            for candidate in output_dir.glob(f"{candidate_name}-*.*"):
                candidate_suffix = candidate.stem.removeprefix(name_prefix)
                if len(candidate_suffix) == 4 and candidate_suffix.isdigit():
                    period_candidates.add(candidate)
        sorted_candidates = sorted(period_candidates)
        if len(sorted_candidates) == 1:
            return sorted_candidates[0]
        if len(sorted_candidates) > 1:
            candidate_text = ", ".join(path.name for path in sorted_candidates)
            joined = " / ".join(report_names)
            raise RuntimeError(f"导出文件匹配不唯一：{joined}（候选：{candidate_text}）")

    joined = " / ".join(report_names)
    raise FileNotFoundError(f"未找到导出文件：{joined}-{suffix}.*（目录：{output_dir}）")


def stream_subprocess(command: list[str], cwd: Path, log: Callable[[str], None], prefix: str) -> None:
    recent_output: list[str] = []
    process = subprocess.Popen(
        command,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    assert process.stdout is not None
    for line in process.stdout:
        text = line.rstrip()
        if text:
            recent_output.append(text)
            recent_output = recent_output[-12:]
            log(f"[{prefix}] {text}")
    exit_code = process.wait()
    if exit_code != 0:
        detail = ""
        if recent_output:
            detail = "；最近输出：" + " | ".join(recent_output[-4:])
        raise RuntimeError(f"{prefix} 执行失败，退出码：{exit_code}{detail}")


def run_fetch_task(
    task: FetchTask,
    *,
    business_date: date,
    runtime_root: Path,
    log: Callable[[str], None],
    headless: bool,
    username: str | None,
    password: str | None,
    chrome_path: str | None,
    max_attempts: int = 2,
) -> Path:
    output_dir = build_task_output_dir(task, runtime_root)
    output_dir.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable,
        task.script_path.name,
        "--business-date",
        format_business_date(business_date),
        "--output-dir",
        str(output_dir.parent),
        "--output-folder-name",
        output_dir.name,
        "--report-keys",
        ",".join(task.report_keys),
    ]
    command.extend(task.extra_args)
    command.append("--headless" if headless else "--headed")
    if username:
        command.extend(["--username", username])
    if password:
        command.extend(["--password", password])
    if chrome_path:
        command.extend(["--chrome-path", chrome_path])

    max_attempts = max(1, int(max_attempts))
    for attempt_index in range(1, max_attempts + 1):
        suffix = "" if attempt_index == 1 else f"（重试 {attempt_index - 1}/{max_attempts - 1}）"
        log(f"开始抓取：{task.label}{suffix}")
        try:
            stream_subprocess(command, task.script_path.parent, log, task.label)
            # ponytail: one configured report writes one Excel file into this run's unique directory.
            export_count = sum(
                1
                for path in output_dir.iterdir()
                if path.is_file() and path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}
            )
            if export_count < len(task.report_keys):
                raise RuntimeError(
                    f"{task.label} 导出结果不完整：预期 {len(task.report_keys)} 个 Excel，实际 {export_count} 个"
                )
            break
        except RuntimeError as exc:
            if UPSTREAM_DATE_NOT_READY_MARKER in str(exc):
                raise
            if attempt_index >= max_attempts:
                raise
            log(f"{task.label} 本次抓取失败，准备重试。原因：{exc}")
            time.sleep(5)
    log(f"抓取完成：{task.label}")
    return output_dir


def clear_worksheet(target_ws) -> None:
    merged_ranges = [str(item) for item in target_ws.merged_cells.ranges]
    for merged in merged_ranges:
        target_ws.unmerge_cells(merged)
    for row in target_ws.iter_rows():
        for cell in row:
            cell.value = None
            cell._style = copy(cell._style)
    target_ws.row_dimensions.clear()
    target_ws.column_dimensions.clear()


def copy_worksheet_contents(source_ws, target_ws) -> None:
    clear_worksheet(target_ws)
    for column_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_letter] = copy(dimension)
    for row_index, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_index] = copy(dimension)

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)

    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))


def replace_workbook_sheets(
    workbook_path: Path,
    export_paths: dict[str, Path],
    mappings: tuple[SheetUpdateMapping, ...],
    *,
    log: Callable[[str], None],
    keep_vba: bool = False,
) -> None:
    workbook = load_workbook(workbook_path, keep_vba=keep_vba)
    sources = {
        mapping.target_sheet: load_workbook(export_paths[mapping.target_sheet], data_only=False)
        for mapping in mappings
    }
    try:
        for mapping in mappings:
            target_ws = workbook[mapping.target_sheet]
            source_wb = sources[mapping.target_sheet]
            source_ws = source_wb[source_wb.sheetnames[0]]
            log(f"回填工作表：{mapping.target_sheet} <- {source_ws.title}")
            copy_worksheet_contents(source_ws, target_ws)

        temp_path = workbook_path.with_name(f"{workbook_path.stem}.updating{workbook_path.suffix}")
        workbook.save(temp_path)
        temp_path.replace(workbook_path)
    finally:
        safe_close_workbook(workbook)
        for source in sources.values():
            safe_close_workbook(source)


def rebuild_dashboard(
    *,
    business_date: date,
    leads_path: Path,
    arrival_path: Path,
    out_path: Path,
    summary_path: Path,
    log: Callable[[str], None],
    archive_root: Path = MONTHLY_ARCHIVE_DIR,
    archive_index_path: Path = MONTHLY_ARCHIVE_INDEX,
    docs_root: Path = DOCS_DIR,
) -> dict[str, str | bool]:
    payload = build_payload(leads_path, arrival_path, report_date_override=business_date)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    changed, existing_payload = write_json_if_changed(
        out_path,
        payload,
        encoding="utf-8",
        volatile_field_paths=(("meta", "generatedAt"),),
    )
    if not changed and existing_payload is not None:
        payload = existing_payload

    archive_seed_summary = build_run_summary(payload, leads_path, arrival_path, out_path, summary_path, changed)
    archive_info = write_monthly_archive(
        payload,
        archive_seed_summary,
        archive_root=archive_root,
        index_path=archive_index_path,
        docs_root=docs_root,
    )
    summary_payload = build_run_summary(payload, leads_path, arrival_path, out_path, summary_path, changed, archive_info)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_changed, _ = write_json_if_changed(
        summary_path,
        summary_payload,
        encoding="utf-8",
        volatile_field_paths=(("generatedAt",),),
    )
    log(f"dashboard.json {'updated' if changed else 'unchanged'}: {out_path}")
    log(f"dashboard.summary.json {'updated' if summary_changed else 'unchanged'}: {summary_path}")
    log(f"dashboard archive updated: {archive_info.get('monthKey')}")
    return {
        "dashboardChanged": changed,
        "summaryChanged": summary_changed,
        "archiveDashboardChanged": bool(archive_info.get("dashboardChanged")),
        "archiveSummaryChanged": bool(archive_info.get("summaryChanged")),
        "archiveIndexChanged": bool(archive_info.get("indexChanged")),
    }


def build_runtime_dir(business_date: date) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return RUNTIME_ROOT / f"{business_date.strftime('%Y%m%d')}_{timestamp}"


def run_update(
    *,
    business_date: date | None = None,
    leads_path: Path = LEADS_BOOK,
    arrival_path: Path = ARRIVAL_BOOK,
    out_path: Path = OUT_JSON,
    summary_path: Path = SUMMARY_JSON,
    log: Callable[[str], None] = default_log,
    headless: bool = True,
    username: str | None = None,
    password: str | None = None,
    chrome_path: str | None = None,
    max_attempts: int = 2,
    keep_runtime: bool = False,
    leads_only: bool = False,
) -> dict[str, object]:
    resolved_business_date = business_date or parse_business_date()
    runtime_dir = build_runtime_dir(resolved_business_date)
    runtime_dir.mkdir(parents=True, exist_ok=True)
    log(f"本次业务日期：{format_business_date(resolved_business_date)}")
    log("更新范围：仅线索 4 张表" if leads_only else "更新范围：线索与来店共 10 张表")
    log(f"运行目录：{runtime_dir}")

    fetch_tasks = LEADS_FETCH_TASKS if leads_only else FETCH_TASKS
    sheet_mappings = LEADS_SHEET_MAPPINGS if leads_only else SHEET_MAPPINGS
    export_paths: dict[str, Path] = {}
    succeeded = False
    try:
        for task in fetch_tasks:
            output_dir = run_fetch_task(
                task,
                business_date=resolved_business_date,
                runtime_root=runtime_dir,
                log=log,
                headless=headless,
                username=username,
                password=password,
                chrome_path=chrome_path,
                max_attempts=max_attempts,
            )
            for mapping in sheet_mappings:
                if mapping.target_sheet in export_paths:
                    continue
                try:
                    export_paths[mapping.target_sheet] = resolve_export_path(
                        output_dir,
                        mapping.export_names,
                        resolved_business_date,
                        allow_period_suffix=mapping.allow_period_suffix,
                    )
                except FileNotFoundError:
                    continue

        missing_reports = [
            mapping.result_label
            for mapping in sheet_mappings
            if mapping.target_sheet not in export_paths
        ]
        if missing_reports:
            raise RuntimeError(f"缺少导出结果：{', '.join(missing_reports)}")

        replace_workbook_sheets(
            leads_path,
            export_paths,
            LEADS_SHEET_MAPPINGS,
            log=log,
            keep_vba=True,
        )
        if not leads_only:
            replace_workbook_sheets(
                arrival_path,
                export_paths,
                ARRIVAL_SHEET_MAPPINGS,
                log=log,
                keep_vba=arrival_path.suffix.lower() == ".xlsm",
            )
        rebuild_summary = rebuild_dashboard(
            business_date=resolved_business_date,
            leads_path=leads_path,
            arrival_path=arrival_path,
            out_path=out_path,
            summary_path=summary_path,
            log=log,
        )
        result = {
            "businessDate": format_business_date(resolved_business_date),
            "updateScope": "leads" if leads_only else "all",
            "runtimeDir": str(runtime_dir),
            "exports": {
                mapping.result_label: str(export_paths[mapping.target_sheet])
                for mapping in sheet_mappings
            },
            **rebuild_summary,
        }
        log("更新流程完成。")
        succeeded = True
        return result
    finally:
        if succeeded and not keep_runtime and runtime_dir.exists():
            shutil.rmtree(runtime_dir, ignore_errors=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch daily browser reports and update AI Digest workbooks.")
    parser.add_argument("--business-date", default="", help="业务日期，默认取今天的 N-1，支持 YYYY-MM-DD / YYYYMMDD")
    parser.add_argument("--username", default="", help="可选：覆盖默认登录账号")
    parser.add_argument("--password", default="", help="可选：覆盖默认登录密码")
    parser.add_argument("--chrome-path", default="", help="可选：指定 Chrome 路径")
    parser.add_argument("--max-attempts", type=int, default=2, help="每个抓取任务失败后的最大尝试次数，默认 2")
    parser.add_argument("--headed", action="store_true", help="启用有头模式，便于调试")
    parser.add_argument("--keep-runtime", action="store_true", help="保留运行时导出与 trace 目录")
    parser.add_argument(
        "--leads-only",
        action="store_true",
        help="仅抓取并回填 4 张线索表，不更新来店工作簿",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = run_update(
        business_date=parse_business_date(args.business_date or None),
        log=default_log,
        headless=not args.headed,
        username=args.username or None,
        password=args.password or None,
        chrome_path=args.chrome_path or None,
        max_attempts=args.max_attempts,
        keep_runtime=args.keep_runtime,
        leads_only=args.leads_only,
    )
    print(result, flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
