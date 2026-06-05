from __future__ import annotations

import argparse
import calendar
import json
from collections import defaultdict
from functools import lru_cache
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = ROOT / "docs"
LEADS_BOOK = ROOT / "data" / "source" / "NEV+ICE_xsai.xlsm"
ARRIVAL_BOOK = ROOT / "data" / "source" / "NEV+ICE_ldai.xlsx"
OUT_JSON = ROOT / "docs" / "data" / "dashboard.json"
SUMMARY_JSON = ROOT / "docs" / "data" / "dashboard.summary.json"
MONTHLY_ARCHIVE_DIR = ROOT / "docs" / "data" / "monthly"
MONTHLY_ARCHIVE_INDEX = MONTHLY_ARCHIVE_DIR / "index.json"

NEV_MODELS = [
    ("nx8", "NX8", "NX8"),
    ("n7", "N7", "N7"),
    ("n6", "N6", "N6"),
    ("TEANA-harmony", "天籁·鸿蒙座舱", "天籁·鸿蒙座舱"),
]
NEV_BRIEF_ORDER = ["NX8", "N7", "N6", "天籁·鸿蒙座舱"]
BRIEF_MARKERS = ["①", "②", "③", "④", "⑤", "⑥"]
SYLPHY_TARGET_OVERRIDES = {
    (2026, 4): [
        1827, 1828, 1828, 1772, 1772, 1772, 1828, 1828, 1828, 1828,
        1773, 1773, 1828, 1828, 1828, 1828, 1828, 1772, 1771, 1828,
        1828, 1828, 1830, 1830, 1771, 1771, 1830, 1830, 1831, 1831,
    ],
    (2026, 5): [
        1622, 1622, 1622, 1622, 1621, 1409, 1408, 1408, 1408, 1620,
        1407, 1407, 1407, 1407, 1407, 1619, 1619, 1407, 1406, 1405,
        1405, 1405, 1617, 1617, 1404, 1404, 1403, 1401, 1401, 1613,
        1613,
    ],
    (2026, 6): [
        1158, 1158, 1158, 1159, 1159, 1335, 1335, 1160, 1161, 1164,
        1164, 1164, 1339, 1339, 1164, 1164, 1164, 1164, 1339, 1339,
        1339, 1165, 1165, 1165, 1166, 1166, 1341, 1342, 1168, 1168,
    ],
}
AGGREGATE_FIELDS = ("newLeads", "validLeads", "storeLeads", "arrivals", "leads", "orders", "deals")
REQUIRED_LEADS_SHEETS = ("参数", "目标竖版", "全国按日NEV", "全国按日ICE", "十五代轩逸按日")
REQUIRED_ARRIVAL_SHEETS = ("NEV本期来店", "NEV同期来店", "ICE本期来店", "ICE同期来店")
REQUIRED_HEADERS = {
    "目标竖版": (2, ("合计",)),
    "全国按日NEV": (2, ("新增线索量", "有效线索量", "门店线索总量", "新增到店量")),
    "全国按日ICE": (1, ("按日", "线索总量", "有效线索量", "到店量", "订单量", "成交量")),
    "十五代轩逸按日": (1, ("按日", "线索总量", "有效线索量", "到店量", "订单量", "成交量")),
}

SPECIAL_DAY_OFFS = {
    date(2026, 4, 6),
}

OFFICIAL_HOLIDAY_RULES: dict[int, dict[str, set[date]]] = {
    2026: {
        "holidays": {
            date(2026, 1, 1), date(2026, 1, 2), date(2026, 1, 3),
            date(2026, 2, 15), date(2026, 2, 16), date(2026, 2, 17), date(2026, 2, 18),
            date(2026, 2, 19), date(2026, 2, 20), date(2026, 2, 21), date(2026, 2, 22), date(2026, 2, 23),
            date(2026, 4, 5), date(2026, 4, 6),
            date(2026, 5, 1), date(2026, 5, 2), date(2026, 5, 3), date(2026, 5, 4), date(2026, 5, 5),
            date(2026, 6, 19), date(2026, 6, 20), date(2026, 6, 21),
            date(2026, 9, 25), date(2026, 9, 26), date(2026, 9, 27),
            date(2026, 10, 1), date(2026, 10, 2), date(2026, 10, 3), date(2026, 10, 4), date(2026, 10, 5),
            date(2026, 10, 6), date(2026, 10, 7),
        },
        "makeup_workdays": {
            date(2026, 1, 4),
            date(2026, 2, 14),
            date(2026, 2, 28),
            date(2026, 5, 9),
            date(2026, 9, 20),
            date(2026, 10, 10),
        },
    },
}


@lru_cache(maxsize=None)
def month_start(value: date) -> date:
    return value.replace(day=1)


@lru_cache(maxsize=None)
def month_end(value: date) -> date:
    return value.replace(day=calendar.monthrange(value.year, value.month)[1])


def month_key(value: date) -> str:
    return f"{value.year:04d}-{value.month:02d}"


def previous_month(value: date) -> date:
    if value.month == 1:
        return date(value.year - 1, 12, 1)
    return date(value.year, value.month - 1, 1)


@lru_cache(maxsize=None)
def month_dates(value: date) -> list[date]:
    start = month_start(value)
    end = month_end(value)
    return [date(start.year, start.month, day) for day in range(1, end.day + 1)]


@lru_cache(maxsize=None)
def aligned_previous_date(current_date: date) -> date | None:
    prev = previous_month(current_date)
    last = calendar.monthrange(prev.year, prev.month)[1]
    if current_date.day > last:
        return None
    return date(prev.year, prev.month, current_date.day)


@lru_cache(maxsize=None)
def coerce_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%Y年%m月%d日"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


@lru_cache(maxsize=None)
def num(value: Any) -> int | float | None:
    if value in (None, "", "-", "/", "#N/A", "#REF!", "#VALUE!"):
        return None
    if isinstance(value, (int, float)):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return None
        try:
            parsed = float(text)
        except ValueError:
            return None
        return int(parsed) if parsed.is_integer() else parsed
    return None


def normalize_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_path_display(path: Path) -> str:
    text = path.as_posix()
    if len(text) >= 2 and text[1] == ":":
        return text[0].upper() + text[1:]
    return text


def build_docs_data_url(path: Path, *, docs_root: Path = DOCS_DIR) -> str:
    relative = path.relative_to(docs_root).as_posix()
    return f"./{relative}"


def serialize_payload(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def write_text_if_changed(path: Path, content: str, *, encoding: str = "utf-8") -> bool:
    existing = path.read_text(encoding=encoding) if path.exists() else None
    if existing == content:
        return False
    path.write_text(content, encoding=encoding)
    return True


def read_json_file(path: Path, *, encoding: str = "utf-8") -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding=encoding))
    except (OSError, json.JSONDecodeError):
        return None


def without_volatile_fields(payload: dict[str, Any], field_paths: tuple[tuple[str, ...], ...]) -> dict[str, Any]:
    cloned = {**payload}
    for path in field_paths:
        parent: Any = cloned
        for key in path[:-1]:
            if not isinstance(parent, dict) or key not in parent:
                parent = None
                break
            copied = {**parent[key]}
            parent[key] = copied
            parent = copied
        if isinstance(parent, dict):
            parent.pop(path[-1], None)
    return cloned


def write_json_if_changed(
    path: Path,
    payload: dict[str, Any],
    *,
    encoding: str = "utf-8",
    volatile_field_paths: tuple[tuple[str, ...], ...] = (),
) -> tuple[bool, dict[str, Any] | None]:
    existing_payload = read_json_file(path, encoding=encoding)
    if existing_payload is not None:
        if without_volatile_fields(existing_payload, volatile_field_paths) == without_volatile_fields(payload, volatile_field_paths):
            return False, existing_payload
    content = serialize_payload(payload)
    changed = write_text_if_changed(path, content, encoding=encoding)
    return changed, existing_payload


def file_mtime_iso(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")


def validate_workbook_sheets(workbook: Workbook, required_sheets: tuple[str, ...], workbook_label: str) -> None:
    missing = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"{workbook_label} 缺少必需工作表：{', '.join(missing)}")


def validate_sheet_headers(ws, row_index: int, required_headers: tuple[str, ...], sheet_name: str) -> None:
    headers = header_map(ws, row_index)
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"工作表 '{sheet_name}' 缺少必需列：{', '.join(missing)}")


def validate_workbook_structure(leads: Workbook, arrival: Workbook) -> None:
    validate_workbook_sheets(leads, REQUIRED_LEADS_SHEETS, "线索工作簿")
    validate_workbook_sheets(arrival, REQUIRED_ARRIVAL_SHEETS, "来店工作簿")
    for sheet_name, (row_index, headers) in REQUIRED_HEADERS.items():
        validate_sheet_headers(leads[sheet_name], row_index, headers, sheet_name)


def resolve_report_date_fallback() -> date | None:
    summary_payload = read_json_file(SUMMARY_JSON)
    if isinstance(summary_payload, dict):
        report_date = coerce_date(summary_payload.get("reportDate"))
        if report_date is not None:
            return report_date

    dashboard_payload = read_json_file(OUT_JSON)
    if isinstance(dashboard_payload, dict):
        meta = dashboard_payload.get("meta")
        if isinstance(meta, dict):
            report_date = coerce_date(meta.get("reportDate"))
            if report_date is not None:
                return report_date
    return None


def validate_report_date_cell(leads: Workbook) -> date:
    report_date = coerce_date(leads["参数"]["C2"].value)
    if report_date is not None:
        return report_date

    fallback = resolve_report_date_fallback()
    if fallback is not None:
        return fallback
    raise ValueError("参数!C2 未读取到有效日期，且未找到可用的报表日期回退值。")


def ratio(current: int | float | None, target: int | float | None) -> float | None:
    if current is None or target in (None, 0):
        return None
    return float(current) / float(target)


def delta_ratio(current: int | float | None, previous: int | float | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return (float(current) - float(previous)) / float(previous)


def fmt_count(value: int | float | None, fallback: str = "-") -> str:
    return fallback if value is None else f"{int(round(float(value))):,}"


def fmt_plain(value: int | float | None, fallback: str = "-") -> str:
    return fallback if value is None else str(int(round(float(value))))


def fmt_percent(value: int | float | None, fallback: str = "-") -> str:
    return fallback if value is None else f"{float(value):.1%}"


def fmt_axis_date(value: date | None) -> str:
    return "-" if value is None else f"{value.month}/{value.day}"


def fmt_sheet_date(value: date | None) -> str:
    return "-" if value is None else f"{value.year}/{value.month}/{value.day}"


@lru_cache(maxsize=None)
def get_day_calendar_meta(value: date | None) -> dict[str, bool | str]:
    if value is None:
        return {
            "dayType": "none",
            "isHoliday": False,
            "isWeekend": False,
            "isMakeupWorkday": False,
            "isRegularWorkday": False,
            "highlight": False,
        }

    year_rules = OFFICIAL_HOLIDAY_RULES.get(value.year, {})
    holidays = year_rules.get("holidays", set())
    makeup_workdays = year_rules.get("makeup_workdays", set())
    is_holiday = value in holidays or value in SPECIAL_DAY_OFFS
    is_makeup_workday = value in makeup_workdays
    is_weekend = value.weekday() >= 5 and not is_holiday and not is_makeup_workday
    is_regular_workday = not is_holiday and not is_weekend and not is_makeup_workday

    if is_holiday:
        day_type = "holiday"
    elif is_makeup_workday:
        day_type = "makeupWorkday"
    elif is_weekend:
        day_type = "weekend"
    else:
        day_type = "regularWorkday"

    return {
        "dayType": day_type,
        "isHoliday": is_holiday,
        "isWeekend": is_weekend,
        "isMakeupWorkday": is_makeup_workday,
        "isRegularWorkday": is_regular_workday,
        "highlight": is_holiday or is_weekend or is_makeup_workday,
    }


def is_day_off(value: date | None) -> bool:
    if value is None:
        return False
    return bool(get_day_calendar_meta(value)["isHoliday"])


def build_single_day_meta(prefix: str, value: date | None) -> dict[str, Any]:
    meta = get_day_calendar_meta(value)
    cap_prefix = prefix[0].upper() + prefix[1:]
    return {
        f"{prefix}Date": value.isoformat() if value else None,
        f"{prefix}DayType": meta["dayType"],
        f"highlight{cap_prefix}": meta["highlight"],
        f"is{cap_prefix}Holiday": meta["isHoliday"],
        f"is{cap_prefix}Weekend": meta["isWeekend"],
        f"is{cap_prefix}MakeupWorkday": meta["isMakeupWorkday"],
        f"is{cap_prefix}RegularWorkday": meta["isRegularWorkday"],
    }


def build_column_calendar_meta(current_date: date | None, previous_date: date | None) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    meta.update(build_single_day_meta("current", current_date))
    meta.update(build_single_day_meta("previous", previous_date))
    return meta


def nice_axis_max(values: list[int | float | None]) -> int:
    nums = [float(v) for v in values if isinstance(v, (int, float))]
    if not nums:
        return 1
    raw = max(nums)
    if raw <= 5000:
        step = 1000
    elif raw <= 20000:
        step = 5000
    elif raw <= 100000:
        step = 10000
    else:
        step = 50000
    return int(((raw + step - 1) // step) * step)


def header_map(ws, row_index: int) -> dict[str, int]:
    return {
        str(ws.cell(row_index, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row_index, col).value not in (None, "")
    }


def load_nev_targets(ws, start_date: date, end_date: date) -> dict[str, dict[date, int | float]]:
    headers = header_map(ws, 2)
    total_col = headers["合计"]
    result: dict[str, dict[date, int | float]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=3, values_only=True):
        current_date = coerce_date(row[0])
        model = row[1]
        total = num(row[total_col - 1])
        if current_date and model and total is not None and start_date <= current_date <= end_date:
            result[str(model).strip()][current_date] = total
    return result


def load_nev_daily(ws, start_date: date, end_date: date) -> dict[str, dict[date, dict[str, int | float | None]]]:
    headers = header_map(ws, 2)
    result: dict[str, dict[date, dict[str, int | float | None]]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=4, values_only=True):
        current_date = coerce_date(row[2])
        model = row[3]
        if current_date is None or not model or not (start_date <= current_date <= end_date):
            continue
        result[str(model).strip()][current_date] = {
            "newLeads": num(row[headers["新增线索量"] - 1]),
            "validLeads": num(row[headers["有效线索量"] - 1]),
            "storeLeads": num(row[headers["门店线索总量"] - 1]),
            "arrivals": num(row[headers["新增到店量"] - 1]),
        }
    return result


def load_ice_daily(ws, start_date: date, end_date: date) -> dict[date, dict[str, int | float | None]]:
    headers = header_map(ws, 1)
    result: dict[date, dict[str, int | float | None]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        current_date = coerce_date(row[headers["按日"] - 1])
        if current_date is None or not (start_date <= current_date <= end_date):
            continue
        leads = num(row[headers["线索总量"] - 1])
        valid_leads = num(row[headers["有效线索量"] - 1])
        arrivals = num(row[headers["到店量"] - 1])
        orders = num(row[headers["订单量"] - 1])
        deals = num(row[headers["成交量"] - 1])
        result[current_date] = {
            "leads": leads,
            "validLeads": valid_leads,
            "arrivals": arrivals,
            "orders": orders,
            "deals": deals,
        }
    return result


def aggregate_daily_series(*groups: dict[date, dict[str, int | float | None]]) -> dict[date, dict[str, int | float | None]]:
    merged: dict[date, dict[str, int | float | None]] = {}
    all_dates = sorted({d for group in groups for d in group})
    for current_date in all_dates:
        totals = {key: 0 for key in AGGREGATE_FIELDS}
        for group in groups:
            item = group.get(current_date)
            if not item:
                continue
            for key in AGGREGATE_FIELDS:
                value = item.get(key)
                if isinstance(value, (int, float)):
                    totals[key] += value
        merged[current_date] = {key: (value or None) for key, value in totals.items()}
    return merged


def aggregate_targets(*groups: dict[date, int | float]) -> dict[date, int | float]:
    merged: dict[date, int | float] = {}
    all_dates = sorted({d for group in groups for d in group})
    for current_date in all_dates:
        values = [group[current_date] for group in groups if current_date in group]
        if values:
            merged[current_date] = sum(values)
    return merged


def build_running_totals(values: list[int | float | None], *, stop_at: int | None = None) -> list[int | float | None]:
    totals: list[int | float | None] = []
    running = 0
    for index, value in enumerate(values):
        if stop_at is not None and index > stop_at:
            totals.append(None)
            continue
        if isinstance(value, (int, float)):
            running += value
        totals.append(running)
    return totals


def build_column_meta(current_date: date, previous_date: date | None) -> dict[str, Any]:
    return build_column_calendar_meta(current_date, previous_date)


def build_monthly_series_context(
    report_date: date,
    current_actuals: dict[date, dict[str, int | float | None]],
    previous_actuals: dict[date, dict[str, int | float | None]],
    actual_key: str,
    current_targets: dict[date, int | float] | None = None,
) -> dict[str, Any]:
    dates = month_dates(report_date)
    report_index = report_date.day - 1
    prev_dates: list[str] = []
    curr_dates: list[str] = []
    prev_daily: list[int | float | None] = []
    target_daily: list[int | float | None] = []
    curr_daily: list[int | float | None] = []
    chart_daily: list[int | float | None] = []
    day_delta: list[float | None] = []
    column_meta: list[dict[str, Any]] = []
    prev_run = 0
    target_run = 0
    curr_run = 0
    has_target = bool(current_targets) and any(value is not None for value in current_targets.values())
    prev_cum: list[int | float | None] = []
    target_cum: list[int | float | None] = []
    curr_cum: list[int | float | None] = []

    for current_date in dates:
        previous_date = aligned_previous_date(current_date)
        prev_value = previous_actuals.get(previous_date, {}).get(actual_key) if previous_date else None
        target_value = current_targets.get(current_date) if current_targets else None
        raw_actual = current_actuals.get(current_date, {}).get(actual_key)
        curr_value = raw_actual if current_date <= report_date else None
        chart_value = raw_actual if raw_actual is not None and current_date <= report_date else 0

        prev_dates.append(fmt_sheet_date(previous_date))
        curr_dates.append(fmt_sheet_date(current_date))
        prev_daily.append(prev_value)
        target_daily.append(target_value)
        curr_daily.append(curr_value)
        chart_daily.append(chart_value)

        if previous_date is not None:
            if isinstance(prev_value, (int, float)):
                prev_run += prev_value
            prev_cum.append(prev_run)
        else:
            prev_cum.append(None)

        if has_target:
            if isinstance(target_value, (int, float)):
                target_run += target_value
            target_cum.append(target_run)
        else:
            target_cum.append(None)

        if current_date <= report_date:
            if isinstance(curr_value, (int, float)):
                curr_run += curr_value
            curr_cum.append(curr_run)
        else:
            curr_cum.append(None)

        day_delta.append(delta_ratio(curr_value, prev_value))
        column_meta.append(build_column_meta(current_date, previous_date))

    return {
        "dates": dates,
        "reportIndex": report_index,
        "prevDates": prev_dates,
        "currDates": curr_dates,
        "prevDaily": prev_daily,
        "targetDaily": target_daily,
        "currDaily": curr_daily,
        "chartDaily": chart_daily,
        "prevCumulative": prev_cum,
        "targetCumulative": target_cum,
        "currCumulative": curr_cum,
        "dayDelta": day_delta,
        "columnMeta": column_meta,
        "hasTarget": has_target,
    }


def build_valid_leads_control_trend(
    report_date: date,
    current_actuals: dict[date, dict[str, int | float | None]],
    previous_actuals: dict[date, dict[str, int | float | None]],
) -> dict[str, Any]:
    context = build_monthly_series_context(report_date, current_actuals, previous_actuals, "validLeads")
    dates = context["dates"]
    report_index = context["reportIndex"]
    prev_dates = context["prevDates"]
    curr_dates = context["currDates"]
    prev_daily = context["prevDaily"]
    curr_daily = context["currDaily"]
    chart_daily = context["chartDaily"]
    prev_cum = context["prevCumulative"]
    curr_cum = context["currCumulative"]
    day_delta = context["dayDelta"]
    column_meta = context["columnMeta"]
    cumulative_delta = [delta_ratio(curr_value, prev_value) for curr_value, prev_value in zip(curr_cum, prev_cum)]
    daily_actual = curr_daily[report_index]
    daily_previous = prev_daily[report_index]
    cumulative_actual = curr_cum[report_index]
    cumulative_previous = prev_cum[report_index]

    return {
        "chartTitle": f"{report_date.month}月全车系有效线索趋势",
        "summary": {
            "items": [
                {"label": "累计本期实绩", "value": normalize_scalar(cumulative_actual), "displayValue": fmt_count(cumulative_actual)},
                {"label": "累计上期实绩", "value": normalize_scalar(cumulative_previous), "displayValue": fmt_count(cumulative_previous)},
                {
                    "label": "累计环比",
                    "value": normalize_scalar(delta_ratio(cumulative_actual, cumulative_previous)),
                    "displayValue": fmt_percent(delta_ratio(cumulative_actual, cumulative_previous)),
                },
                {"label": "当日本期实绩", "value": normalize_scalar(daily_actual), "displayValue": fmt_count(daily_actual)},
                {"label": "当日上期实绩", "value": normalize_scalar(daily_previous), "displayValue": fmt_count(daily_previous)},
                {
                    "label": "当日环比",
                    "value": normalize_scalar(delta_ratio(daily_actual, daily_previous)),
                    "displayValue": fmt_percent(delta_ratio(daily_actual, daily_previous)),
                },
            ]
        },
        "matrix": {
            "labels": [fmt_axis_date(item) for item in dates],
            "columnMeta": column_meta,
            "visibleRowKeys": [
                "previousActual",
                "previousCumulative",
                "actual",
                "cumulativeActual",
                "dayDelta",
                "cumulativeDelta",
            ],
            "rows": [
                {"key": "previousDate", "label": "上期日期", "displayValues": prev_dates},
                {"key": "currentDate", "label": "本期日期", "displayValues": curr_dates},
                {"key": "previousActual", "label": "上期实绩", "displayValues": [fmt_plain(v) for v in prev_daily]},
                {"key": "previousCumulative", "label": "上期累计实绩", "displayValues": [fmt_plain(v) for v in prev_cum]},
                {"key": "actual", "label": "本期实绩", "displayValues": [fmt_plain(v) for v in curr_daily]},
                {"key": "cumulativeActual", "label": "本期累计实绩", "displayValues": [fmt_plain(v) for v in curr_cum]},
                {"key": "dayDelta", "label": "环比", "displayValues": [fmt_percent(v) for v in day_delta]},
                {"key": "cumulativeDelta", "label": "累计环比", "displayValues": [fmt_percent(v) for v in cumulative_delta]},
            ],
        },
        "chart": {
            "labels": [fmt_axis_date(item) for item in dates],
            "reportDayIndex": report_index,
            "dailyAxisMax": nice_axis_max([*prev_daily, *chart_daily]),
            "cumulativeAxisMax": nice_axis_max([*prev_cum, *curr_cum]),
            "seriesDefinitions": [
                {
                    "key": "previousActual",
                    "label": "上期实绩",
                    "type": "bar",
                    "color": "#8da1b8",
                    "fill": "rgba(255,255,255,0.85)",
                    "stroke": "#8da1b8",
                    "strokeWidth": 1.4,
                },
                {"key": "actual", "label": "本期实绩", "type": "bar", "color": "#c20f2f", "fill": "#c20f2f"},
                {"key": "previousCumulative", "label": "上期累计实绩", "type": "line", "color": "#b6c2d0", "strokeWidth": "3"},
                {
                    "key": "cumulativeActual",
                    "label": "本期累计实绩",
                    "type": "line",
                    "color": "#c20f2f",
                    "strokeWidth": "3.5",
                    "markers": True,
                    "markerFill": "#ffffff",
                    "markerStroke": "#c20f2f",
                    "markerRadius": 4.8,
                },
            ],
            "series": {
                "previousActual": [normalize_scalar(v) for v in prev_daily],
                "actual": [normalize_scalar(v) for v in chart_daily],
                "previousCumulative": [normalize_scalar(v) for v in prev_cum],
                "cumulativeActual": [normalize_scalar(v) for v in curr_cum],
            },
            "note": "",
        },
    }


def build_monthly_trend(
    title: str,
    metric_label: str,
    report_date: date,
    current_actuals: dict[date, dict[str, int | float | None]],
    previous_actuals: dict[date, dict[str, int | float | None]],
    actual_key: str,
    current_targets: dict[date, int | float] | None,
) -> dict[str, Any]:
    context = build_monthly_series_context(report_date, current_actuals, previous_actuals, actual_key, current_targets)
    chart_title = build_monthly_chart_title(title, metric_label)
    dates = context["dates"]
    report_index = context["reportIndex"]
    prev_dates = context["prevDates"]
    curr_dates = context["currDates"]
    prev_daily = context["prevDaily"]
    target_daily = context["targetDaily"]
    curr_daily = context["currDaily"]
    chart_daily = context["chartDaily"]
    prev_cum = context["prevCumulative"]
    target_cum = context["targetCumulative"]
    curr_cum = context["currCumulative"]
    day_delta = context["dayDelta"]
    column_meta = context["columnMeta"]
    has_target = context["hasTarget"]
    cumulative_target = target_cum[report_index] if has_target else None
    cumulative_actual = curr_cum[report_index]
    previous_cumulative = prev_cum[report_index]
    daily_target = target_daily[report_index] if has_target else None
    daily_actual = curr_daily[report_index]
    return {
        "chartTitle": f"{report_date.month} 月{chart_title}趋势",
        "summary": {
            "items": [
                {"label": "累计目标", "value": normalize_scalar(cumulative_target), "displayValue": fmt_count(cumulative_target)},
                {"label": "累计实绩", "value": normalize_scalar(cumulative_actual), "displayValue": fmt_count(cumulative_actual)},
                {"label": "累计达成", "value": normalize_scalar(ratio(cumulative_actual, cumulative_target)), "displayValue": fmt_percent(ratio(cumulative_actual, cumulative_target)), "note": f"累计环比 {fmt_percent(delta_ratio(cumulative_actual, previous_cumulative))}"},
                {"label": "当日目标", "value": normalize_scalar(daily_target), "displayValue": fmt_count(daily_target)},
                {"label": "当日实绩", "value": normalize_scalar(daily_actual), "displayValue": fmt_count(daily_actual)},
                {"label": "当日达成", "value": normalize_scalar(ratio(daily_actual, daily_target)), "displayValue": fmt_percent(ratio(daily_actual, daily_target))},
            ]
        },
        "matrix": {
            "labels": [fmt_axis_date(item) for item in dates],
            "columnMeta": column_meta,
            "rows": [
                {"key": "previousDate", "label": "上期日期", "displayValues": prev_dates},
                {"key": "currentDate", "label": "本期日期", "displayValues": curr_dates},
                {"key": "previousActual", "label": "上期实绩", "displayValues": [fmt_plain(v) for v in prev_daily]},
                {"key": "target", "label": "本期目标", "displayValues": [fmt_plain(v) for v in target_daily]},
                {"key": "previousCumulative", "label": "上期累计实绩", "displayValues": [fmt_plain(v) for v in prev_cum]},
                {"key": "actual", "label": "本期实绩", "displayValues": [fmt_plain(v) for v in curr_daily]},
                {"key": "cumulativeTarget", "label": "本期累计目标", "displayValues": [fmt_plain(v) for v in target_cum]},
                {"key": "cumulativeActual", "label": "本期累计实绩", "displayValues": [fmt_plain(v) for v in curr_cum]},
                {"key": "dayDelta", "label": "环比", "displayValues": [fmt_percent(v) for v in day_delta]},
            ],
        },
        "chart": {
            "labels": [fmt_axis_date(item) for item in dates],
            "reportDayIndex": report_index,
            "dailyAxisMax": nice_axis_max([*prev_daily, *target_daily, *chart_daily]),
            "cumulativeAxisMax": nice_axis_max([*prev_cum, *target_cum, *curr_cum]),
            "series": {
                "previousActual": [normalize_scalar(v) for v in prev_daily],
                "target": [normalize_scalar(v) for v in target_daily],
                "actual": [normalize_scalar(v) for v in chart_daily],
                "previousCumulative": [normalize_scalar(v) for v in prev_cum],
                "cumulativeTarget": [normalize_scalar(v) for v in target_cum],
                "cumulativeActual": [normalize_scalar(v) for v in curr_cum],
            },
            "note": "",
        },
    }


def make_card(label: str, value: int | float | None, kind: str = "count", note: str = "") -> dict[str, Any]:
    return {
        "label": label,
        "value": normalize_scalar(value),
        "displayValue": fmt_percent(value) if kind == "percent" else fmt_count(value),
        "note": note,
    }


def build_monthly_chart_title(title: str, metric_label: str) -> str:
    special_titles = {
        ("NEV 总盘", "新增线索"): "NEV 新增线索",
        ("ICE 总盘", "有效线索"): "ICE 有效线索",
    }
    return special_titles.get((title, metric_label), f"{title}{metric_label}")


def build_nev_section(section_id: str, title: str, report_date: date, current_actuals, previous_actuals, current_targets) -> dict[str, Any]:
    latest = current_actuals.get(report_date, {})
    total_new = sum((item.get("newLeads") or 0) for item in current_actuals.values())
    total_target = sum(current_targets.get(item, 0) or 0 for item in month_dates(report_date) if item <= report_date)
    total_arrivals = sum((item.get("arrivals") or 0) for item in current_actuals.values())
    latest_new = latest.get("newLeads")
    latest_target = current_targets.get(report_date)
    return {
        "id": section_id,
        "title": title,
        "headline": "",
        "summary": {
            "cards": [
                make_card("累计新增线索", total_new),
                make_card("累计目标", total_target),
                make_card("累计达成率", ratio(total_new, total_target), kind="percent"),
                make_card("累计新增到店", total_arrivals),
                make_card("当日新增线索", latest_new, note=report_date.isoformat()),
                make_card("当日目标", latest_target),
                make_card("当日达成率", ratio(latest_new, latest_target), kind="percent"),
                make_card("当日新增到店", latest.get("arrivals")),
            ],
            "auxiliary": [],
        },
        "trend": build_monthly_trend(title, "新增线索", report_date, current_actuals, previous_actuals, "newLeads", current_targets),
        "note": "",
        "noteHasError": False,
    }


def build_ice_section(section_id: str, title: str, report_date: date, current_actuals, previous_actuals, current_targets) -> dict[str, Any]:
    latest = current_actuals.get(report_date, {})
    total_leads = sum((item.get("leads") or 0) for item in current_actuals.values())
    total_valid = sum((item.get("validLeads") or 0) for item in current_actuals.values())
    total_arrivals = sum((item.get("arrivals") or 0) for item in current_actuals.values())
    total_orders = sum((item.get("orders") or 0) for item in current_actuals.values())
    total_deals = sum((item.get("deals") or 0) for item in current_actuals.values())
    return {
        "id": section_id,
        "title": title,
        "headline": "",
        "summary": {
            "cards": [
                make_card("累计有效线索", total_valid),
                make_card("累计线索总量", total_leads),
                make_card("累计有效率", ratio(total_valid, total_leads), kind="percent"),
                make_card("累计到店量", total_arrivals),
                make_card("累计订单量", total_orders),
                make_card("累计成交量", total_deals),
                make_card("当日有效线索", latest.get("validLeads"), note=report_date.isoformat()),
                make_card("当日到店量", latest.get("arrivals")),
            ],
            "auxiliary": [],
        },
        "trend": build_monthly_trend(title, "有效线索", report_date, current_actuals, previous_actuals, "validLeads", current_targets),
        "note": "",
        "noteHasError": False,
    }


def build_sylphy_target_series(report_date: date) -> dict[date, int]:
    values = SYLPHY_TARGET_OVERRIDES.get((report_date.year, report_date.month), [])
    return {current_date: values[index] for index, current_date in enumerate(month_dates(report_date)) if index < len(values)}


def build_line_brief(report_date: date, nev_daily, nev_targets, sylphy_daily, sylphy_targets) -> dict[str, Any]:
    total_cum_actual = total_day_actual = total_cum_target = total_day_target = 0
    nev_lines: list[str] = []
    for index, model_name in enumerate(NEV_BRIEF_ORDER):
        actuals = nev_daily.get(model_name, {})
        targets = nev_targets.get(model_name, {})
        cum_actual = sum((item.get("newLeads") or 0) for item in actuals.values())
        day_actual = actuals.get(report_date, {}).get("newLeads")
        cum_target = sum(targets.get(d, 0) or 0 for d in month_dates(report_date) if d <= report_date)
        day_target = targets.get(report_date)
        total_cum_actual += cum_actual
        total_day_actual += day_actual or 0
        total_cum_target += cum_target
        total_day_target += day_target or 0
        marker = BRIEF_MARKERS[index]
        nev_lines.append(f"{marker}{model_name}累计实绩{fmt_count(cum_actual)}，累计达成率{fmt_percent(ratio(cum_actual, cum_target))}；当日实绩{fmt_count(day_actual)}，当日达成率{fmt_percent(ratio(day_actual, day_target))}")
    nev_summary = f"四车累计实绩{fmt_count(total_cum_actual)}，累计达成率{fmt_percent(ratio(total_cum_actual, total_cum_target))}；当日实绩{fmt_count(total_day_actual)}，当日达成率{fmt_percent(ratio(total_day_actual, total_day_target))}"
    sylphy_cum_actual = sum((item.get('validLeads') or 0) for item in sylphy_daily.values())
    sylphy_day_actual = sylphy_daily.get(report_date, {}).get("validLeads")
    sylphy_cum_target = sum(sylphy_targets.get(d, 0) or 0 for d in month_dates(report_date) if d <= report_date)
    sylphy_day_target = sylphy_targets.get(report_date)
    sylphy_line = f"十五代轩逸累计实绩 {fmt_count(sylphy_cum_actual)}，累计达成率 {fmt_percent(ratio(sylphy_cum_actual, sylphy_cum_target))}；当日实绩 {fmt_count(sylphy_day_actual)}，当日达成率 {fmt_percent(ratio(sylphy_day_actual, sylphy_day_target))}"
    headline = f"请查收{report_date.strftime('%m.%d')}线索&来店日报"
    sections = [
        {"kind": "intro", "title": "开场", "lines": ["各位领导：", headline]},
        {"kind": "nev", "title": "NEV线索", "lines": [nev_summary, *nev_lines]},
        {"kind": "sylphy15", "title": "十五代轩逸线索", "lines": [sylphy_line]},
    ]
    return {
        "id": "daily-brief",
        "title": "每日简报",
        "headline": headline,
        "dateLabel": report_date.strftime("%m.%d"),
        "reportDate": report_date.isoformat(),
        "sections": sections,
        "generatedText": "\n\n".join(["\n".join(item["lines"]) if item["kind"] == "intro" else "\n".join([f"【{item['title']}】", *item["lines"]]) for item in sections]),
    }


def build_arrival_brief(report_date: date, arrival_maps: dict[str, dict[date, int | float]]) -> dict[str, Any]:
    total_series = build_arrival_series(report_date, arrival_maps["total_current"], arrival_maps["total_previous"])
    nev_series = build_arrival_series(report_date, arrival_maps["nev_current"], arrival_maps["nev_previous"])
    ice_series = build_arrival_series(report_date, arrival_maps["ice_current"], arrival_maps["ice_previous"])

    total_index = total_series["reportIndex"]
    nev_index = nev_series["reportIndex"]
    ice_index = ice_series["reportIndex"]
    lines = [
        format_arrival_brief_line(
            "全国",
            total_series["currentCumulative"][total_index],
            total_series["previousCumulative"][total_index],
            total_series["currentDaily"][total_index],
            total_series["previousDaily"][total_index],
        ),
        format_arrival_brief_line(
            "NEV",
            nev_series["currentCumulative"][nev_index],
            nev_series["previousCumulative"][nev_index],
            nev_series["currentDaily"][nev_index],
            nev_series["previousDaily"][nev_index],
            marker="①",
        ),
        format_arrival_brief_line(
            "ICE",
            ice_series["currentCumulative"][ice_index],
            ice_series["previousCumulative"][ice_index],
            ice_series["currentDaily"][ice_index],
            ice_series["previousDaily"][ice_index],
            marker="②",
        ),
    ]
    return {
        "kind": "arrival",
        "title": "来店简报",
        "lines": [item for item in lines if item],
        "sourceSheets": ["NEV本期来店", "NEV同期来店", "ICE本期来店", "ICE同期来店"],
    }


def safe_close_workbook(workbook: Any) -> None:
    archive = getattr(workbook, "_archive", None)
    vba_archive = getattr(workbook, "vba_archive", None)
    workbook.close()
    for handle in (archive, vba_archive):
        if handle is not None:
            try:
                handle.close()
            except Exception:
                pass


def aligned_previous_year_date(current_date: date) -> date:
    return date(current_date.year - 1, current_date.month, current_date.day)


def load_arrival_daily_sheet(ws) -> dict[date, int | float]:
    result: dict[date, int | float] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        current_date = coerce_date(row[0])
        current_value = num(row[1])
        if current_date is None or current_value is None:
            continue
        result[current_date] = current_value
    return result


def build_arrival_daily_maps(arrival_wb) -> dict[str, dict[date, int | float]]:
    nev_current = load_arrival_daily_sheet(arrival_wb["NEV本期来店"])
    nev_previous = load_arrival_daily_sheet(arrival_wb["NEV同期来店"])
    ice_current = load_arrival_daily_sheet(arrival_wb["ICE本期来店"])
    ice_previous = load_arrival_daily_sheet(arrival_wb["ICE同期来店"])

    total_current: dict[date, int | float] = {}
    total_previous: dict[date, int | float] = {}
    for current_date in sorted({*nev_current.keys(), *ice_current.keys()}):
        total_current[current_date] = (nev_current.get(current_date) or 0) + (ice_current.get(current_date) or 0)
    for current_date in sorted({*nev_previous.keys(), *ice_previous.keys()}):
        total_previous[current_date] = (nev_previous.get(current_date) or 0) + (ice_previous.get(current_date) or 0)

    return {
        "total_current": total_current,
        "total_previous": total_previous,
        "nev_current": nev_current,
        "nev_previous": nev_previous,
        "ice_current": ice_current,
        "ice_previous": ice_previous,
    }


def build_arrival_series(
    report_date: date,
    current_map: dict[date, int | float],
    previous_map: dict[date, int | float],
) -> dict[str, Any]:
    dates = month_dates(report_date)
    current_daily: list[int | float | None] = []
    previous_daily: list[int | float | None] = []

    for current_date in dates:
        previous_date = aligned_previous_year_date(current_date)
        current_daily.append(current_map.get(current_date) if current_date <= report_date else None)
        previous_daily.append(previous_map.get(previous_date))

    report_indexes = [index for index, value in enumerate(current_daily) if isinstance(value, (int, float))]
    report_index = report_indexes[-1] if report_indexes else max(report_date.day - 1, 0)
    previous_report_indexes = [index for index, value in enumerate(previous_daily) if isinstance(value, (int, float))]
    previous_report_index = previous_report_indexes[-1] if previous_report_indexes else None
    chart_actual = [value if index <= report_index and value is not None else 0 for index, value in enumerate(current_daily)]
    current_cumulative = build_running_totals(current_daily, stop_at=report_index)
    previous_cumulative = (
        build_running_totals(previous_daily, stop_at=previous_report_index)
        if previous_report_index is not None
        else [None] * len(previous_daily)
    )

    return {
        "dates": dates,
        "reportIndex": report_index,
        "previousReportIndex": previous_report_index,
        "currentDaily": current_daily,
        "previousDaily": previous_daily,
        "chartActual": chart_actual,
        "currentCumulative": current_cumulative,
        "previousCumulative": previous_cumulative,
    }


def format_arrival_brief_line(
    label: str,
    current_cumulative: int | float | None,
    previous_cumulative: int | float | None,
    daily_current: int | float | None,
    daily_previous: int | float | None,
    *,
    marker: str = "",
) -> str:
    prefix = f"{marker}{label}"
    return f"{prefix}累计来店{fmt_count(current_cumulative)}；当日来店{fmt_count(daily_current)}"


def build_valid_leads_control_section(
    report_date: date,
    current_actuals: dict[date, dict[str, int | float | None]],
    previous_actuals: dict[date, dict[str, int | float | None]],
) -> dict[str, Any]:
    return {
        "id": "lead-control-total",
        "sectionLabel": "",
        "title": "",
        "headline": "",
        "summary": {"cards": []},
        "trend": build_valid_leads_control_trend(report_date, current_actuals, previous_actuals),
        "note": "",
        "noteHasError": False,
    }


def build_arrival_dashboard(report_date: date, arrival_maps: dict[str, dict[date, int | float]]) -> dict[str, Any]:
    total_series = build_arrival_series(report_date, arrival_maps["total_current"], arrival_maps["total_previous"])
    nev_series = build_arrival_series(report_date, arrival_maps["nev_current"], arrival_maps["nev_previous"])
    ice_series = build_arrival_series(report_date, arrival_maps["ice_current"], arrival_maps["ice_previous"])

    current_dates = total_series["dates"]
    previous_dates = [aligned_previous_year_date(item) for item in current_dates]
    current_daily = total_series["currentDaily"]
    previous_daily = total_series["previousDaily"]
    current_target = [None] * len(current_dates)
    chart_actual = total_series["chartActual"]
    current_cumulative = total_series["currentCumulative"]
    previous_cumulative = total_series["previousCumulative"]
    target_cumulative = [None] * len(current_dates)
    report_index = total_series["reportIndex"]
    month_prefix = f"{current_dates[report_index].month}月" if current_dates and current_dates[report_index] else ""

    matrix_rows = [
        {"key": "previousActual", "label": "同期来店", "displayValues": [fmt_plain(item) for item in previous_daily]},
        {"key": "target", "label": "本期目标", "displayValues": [fmt_plain(item) for item in current_target]},
        {"key": "actual", "label": "本期来店", "displayValues": [fmt_plain(item) for item in current_daily]},
        {"key": "dayDelta", "label": "同比", "displayValues": [fmt_percent(delta_ratio(current, previous)) for current, previous in zip(current_daily, previous_daily)]},
        {"key": "nevActual", "label": "NEV本期实绩", "displayValues": [fmt_plain(item) for item in nev_series["currentDaily"]]},
        {"key": "iceActual", "label": "ICE本期实绩", "displayValues": [fmt_plain(item) for item in ice_series["currentDaily"]]},
        {"key": "iceDelta", "label": "ICE同比", "displayValues": [fmt_percent(delta_ratio(current, previous)) for current, previous in zip(ice_series["currentDaily"], ice_series["previousDaily"])]},
    ]

    trend = {
        "viewType": "arrival",
        "chartTitle": f"{current_dates[report_index].month}月全车系来店日趋势" if current_dates and current_dates[report_index] else "全车系来店日趋势",
        "chartSubtitle": "同期来店 / 本期目标 / 本期来店 / 累计来店",
        "tableTitle": "全国来店明细表",
        "summary": {
            "items": [
                {"label": "累计来店", "value": normalize_scalar(current_cumulative[report_index]), "displayValue": fmt_count(current_cumulative[report_index])},
                {"label": "累计同期来店", "value": normalize_scalar(previous_cumulative[report_index]), "displayValue": fmt_count(previous_cumulative[report_index])},
                {"label": "累计同期同比", "value": normalize_scalar(delta_ratio(current_cumulative[report_index], previous_cumulative[report_index])), "displayValue": fmt_percent(delta_ratio(current_cumulative[report_index], previous_cumulative[report_index]))},
                {"label": "当日来店", "value": normalize_scalar(current_daily[report_index]), "displayValue": fmt_count(current_daily[report_index])},
                {"label": "同期来店", "value": normalize_scalar(previous_daily[report_index]), "displayValue": fmt_count(previous_daily[report_index])},
            ]
        },
        "matrix": {
            "stubLabel": "项目",
            "labels": [fmt_axis_date(item) for item in current_dates],
            "visibleRowKeys": [
                "previousActual",
                "actual",
                "dayDelta",
                "nevActual",
                "iceActual",
                "iceDelta",
            ],
            "columnMeta": [build_column_calendar_meta(current, previous) for current, previous in zip(current_dates, previous_dates)],
            "rows": matrix_rows,
        },
        "chart": {
            "labels": [fmt_axis_date(item) for item in current_dates],
            "reportDayIndex": report_index,
            "dailyAxisMax": nice_axis_max([*previous_daily, *current_target, *chart_actual]),
            "cumulativeAxisMax": nice_axis_max([*previous_cumulative, *target_cumulative, *current_cumulative]),
            "hiddenSeriesKeys": ["target", "cumulativeTarget"],
            "series": {
                "previousActual": [normalize_scalar(item) for item in previous_daily],
                "target": [normalize_scalar(item) for item in current_target],
                "actual": [normalize_scalar(item) for item in chart_actual],
                "previousCumulative": [normalize_scalar(item) for item in previous_cumulative],
                "cumulativeTarget": [normalize_scalar(item) for item in target_cumulative],
                "cumulativeActual": [normalize_scalar(item if index <= report_index else None) for index, item in enumerate(current_cumulative)],
            },
            "seriesDefinitions": [
                {"key": "previousActual", "label": "同期来店", "type": "bar", "color": "#a9c8ff", "fill": "rgba(255,255,255,0.85)", "stroke": "#a9c8ff", "strokeWidth": 1.4},
                {"key": "target", "label": "本期目标", "type": "bar", "color": "#d7d7d7", "fill": "#d7d7d7", "opacity": 0.9},
                {"key": "actual", "label": "本期来店", "type": "bar", "color": "#d40000", "fill": "#d40000"},
                {"key": "cumulativeTarget", "label": f"{month_prefix}累计目标", "type": "line", "color": "#9f9f9f", "dashed": True, "strokeWidth": "3"},
                {"key": "cumulativeActual", "label": f"{month_prefix}累计来店", "type": "line", "color": "#d40000", "strokeWidth": "3.5", "markers": True, "markerFill": "#ffffff", "markerStroke": "#d40000", "markerRadius": 4.8},
                {"key": "previousCumulative", "label": "累计同期来店", "type": "line", "color": "#bfd0ff", "strokeWidth": "3", "markers": False},
            ],
            "note": "",
        },
    }

    section = {
        "id": "arrival-total",
        "sectionLabel": "",
        "title": "",
        "headline": "",
        "summary": {"cards": []},
        "trend": trend,
        "note": "",
        "noteHasError": False,
    }

    return {
        "id": "arrival",
        "pageType": "dashboard",
        "title": "全车系来店日趋势",
        "headline": "",
        "sections": [section],
    }


def build_payload(
    leads_path: Path,
    arrival_path: Path,
    report_date_override: date | None = None,
) -> dict[str, Any]:
    leads = load_workbook(leads_path, data_only=True)
    arrival = load_workbook(arrival_path, data_only=True)
    try:
        validate_workbook_structure(leads, arrival)
        report_date = report_date_override or validate_report_date_cell(leads)

        current_start = month_start(report_date)
        previous_start = previous_month(report_date)
        previous_end = month_end(previous_start)
        current_end = month_end(report_date)

        nev_targets = load_nev_targets(leads["目标竖版"], current_start, current_end)
        nev_daily_all = load_nev_daily(leads["全国按日NEV"], previous_start, current_end)
        ice_daily_all = load_ice_daily(leads["全国按日ICE"], previous_start, current_end)
        sylphy_daily_all = load_ice_daily(leads["十五代轩逸按日"], previous_start, current_end)

        nev_current = {model: {dt: value for dt, value in series.items() if current_start <= dt <= report_date} for model, series in nev_daily_all.items()}
        nev_previous = {model: {dt: value for dt, value in series.items() if previous_start <= dt <= previous_end} for model, series in nev_daily_all.items()}
        ice_current = {dt: value for dt, value in ice_daily_all.items() if current_start <= dt <= report_date}
        ice_previous = {dt: value for dt, value in ice_daily_all.items() if previous_start <= dt <= previous_end}
        sylphy_current = {dt: value for dt, value in sylphy_daily_all.items() if current_start <= dt <= report_date}
        sylphy_previous = {dt: value for dt, value in sylphy_daily_all.items() if previous_start <= dt <= previous_end}
        sylphy_targets = build_sylphy_target_series(report_date)
        nev_total_current = aggregate_daily_series(*(nev_current.get(model_name, {}) for _, _, model_name in NEV_MODELS))
        nev_total_previous = aggregate_daily_series(*(nev_previous.get(model_name, {}) for _, _, model_name in NEV_MODELS))
        arrival_maps = build_arrival_daily_maps(arrival)

        line_brief = build_line_brief(report_date, nev_current, nev_targets, sylphy_current, sylphy_targets)
        arrival_brief = build_arrival_brief(report_date, arrival_maps)
        arrival_dashboard = build_arrival_dashboard(report_date, arrival_maps)

        nev_total_targets = aggregate_targets(*(nev_targets.get(model_name, {}) for _, _, model_name in NEV_MODELS))
        valid_leads_total_current = aggregate_daily_series(nev_total_current, ice_current)
        valid_leads_total_previous = aggregate_daily_series(nev_total_previous, ice_previous)
        data_dates = sorted({*nev_total_current.keys(), *ice_current.keys(), *sylphy_current.keys(), *valid_leads_total_current.keys()})

        dashboards = {
            "brief": {
                "id": "brief",
                "pageType": "brief",
                "title": "每日简报",
                "headline": "",
                "briefing": {**line_brief, "sections": [*line_brief["sections"], arrival_brief], "arrivalBrief": arrival_brief},
                "sections": [],
            },
            "lead-control": {
                "id": "lead-control",
                "pageType": "dashboard",
                "title": "全车系有效线索管控",
                "headline": "",
                "sections": [
                    build_valid_leads_control_section(report_date, valid_leads_total_current, valid_leads_total_previous),
                ],
            },
            "nev": {
                "id": "nev",
                "pageType": "dashboard",
                "title": "NEV 线索趋势",
                "headline": "",
                "sections": [
                    build_nev_section("nev-total", "NEV 总盘", report_date, nev_total_current, nev_total_previous, nev_total_targets),
                    *[
                        build_nev_section(section_id, title, report_date, nev_current.get(model_name, {}), nev_previous.get(model_name, {}), nev_targets.get(model_name, {}))
                        for section_id, title, model_name in NEV_MODELS
                    ],
                ],
            },
            "ice": {
                "id": "ice",
                "pageType": "dashboard",
                "title": "ICE 线索趋势",
                "headline": "",
                "sections": [
                    build_ice_section("ice-total", "ICE 总盘", report_date, ice_current, ice_previous, None),
                    build_ice_section("sylphy-15", "十五代轩逸", report_date, sylphy_current, sylphy_previous, sylphy_targets),
                ],
            },
            "arrival": arrival_dashboard,
        }

        return {
            "meta": {
                "workbook": normalize_path_display(leads_path),
                "workbookName": leads_path.name,
                "arrivalWorkbook": normalize_path_display(arrival_path),
                "arrivalWorkbookName": arrival_path.name,
                "generatedAt": datetime.now().isoformat(timespec="seconds"),
                "workbookModifiedAt": datetime.fromtimestamp(leads_path.stat().st_mtime).isoformat(timespec="seconds"),
                "reportDate": report_date.isoformat(),
                "reportDateLabel": report_date.isoformat(),
                "dataRangeStart": data_dates[0].isoformat() if data_dates else None,
                "dataRangeEnd": data_dates[-1].isoformat() if data_dates else None,
            },
            "analysis": {
                "sheetCount": len(leads.sheetnames),
                "sheetNames": leads.sheetnames,
                "issues": [
                    {"sheet": "每日NEV早报模板", "summary": "线索简报按底层数据重建，避免继续沿用历史模板文案。"},
                    {"sheet": "NEV+ICE_ldai", "summary": "全国来店简报与趋势改为基于 4 张来店底表聚合生成，不再依赖汇总页缓存结果。"},
                ],
            },
            "dashboards": dashboards,
        }
    finally:
        safe_close_workbook(leads)
        safe_close_workbook(arrival)


def build_run_summary(
    payload: dict[str, Any],
    leads_path: Path,
    arrival_path: Path,
    out_path: Path,
    summary_path: Path,
    dashboard_changed: bool,
    archive_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    dashboards = payload.get("dashboards", {})
    analysis = payload.get("analysis", {})
    issues = analysis.get("issues", [])
    meta = payload.get("meta", {})
    outputs = {
        "dashboardJson": normalize_path_display(out_path),
        "dashboardJsonName": out_path.name,
        "dashboardChanged": dashboard_changed,
        "dashboardStatus": "updated" if dashboard_changed else "unchanged",
        "summaryJson": normalize_path_display(summary_path),
        "summaryJsonName": summary_path.name,
    }
    if archive_info:
        outputs.update(
            {
                "archiveMonth": archive_info.get("monthKey"),
                "archiveDashboardJson": archive_info.get("dashboardPath"),
                "archiveSummaryJson": archive_info.get("summaryPath"),
                "archiveIndexJson": archive_info.get("indexPath"),
                "archiveDashboardChanged": archive_info.get("dashboardChanged"),
                "archiveSummaryChanged": archive_info.get("summaryChanged"),
                "archiveIndexChanged": archive_info.get("indexChanged"),
            }
        )
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "reportDate": meta.get("reportDate"),
        "reportDateLabel": meta.get("reportDateLabel"),
        "inputs": {
            "workbook": normalize_path_display(leads_path),
            "workbookName": leads_path.name,
            "workbookModifiedAt": file_mtime_iso(leads_path),
            "arrivalWorkbook": normalize_path_display(arrival_path),
            "arrivalWorkbookName": arrival_path.name,
            "arrivalWorkbookModifiedAt": file_mtime_iso(arrival_path),
        },
        "outputs": outputs,
        "stats": {
            "dashboardCount": len(dashboards),
            "sectionCounts": {
                dashboard_id: len(dashboard.get("sections", []))
                for dashboard_id, dashboard in dashboards.items()
            },
            "sheetCount": analysis.get("sheetCount", 0),
            "issueCount": len(issues),
        },
        "warnings": [
            issue.get("summary")
            for issue in issues
            if isinstance(issue, dict) and issue.get("summary")
        ],
    }


def write_monthly_archive(
    payload: dict[str, Any],
    summary_payload: dict[str, Any],
    *,
    archive_root: Path = MONTHLY_ARCHIVE_DIR,
    index_path: Path = MONTHLY_ARCHIVE_INDEX,
    docs_root: Path = DOCS_DIR,
) -> dict[str, Any]:
    meta = payload.get("meta", {})
    report_date = coerce_date(meta.get("reportDate"))
    if report_date is None:
        raise ValueError("无法为月度归档确定 reportDate。")

    report_month = month_key(report_date)
    archive_dir = archive_root / report_month
    archive_dashboard_path = archive_dir / "dashboard.json"
    archive_summary_path = archive_dir / "dashboard.summary.json"
    archive_dir.mkdir(parents=True, exist_ok=True)

    archive_dashboard_changed, archive_dashboard_existing = write_json_if_changed(
        archive_dashboard_path,
        payload,
        encoding="utf-8",
        volatile_field_paths=(("meta", "generatedAt"),),
    )
    archive_summary_changed, archive_summary_existing = write_json_if_changed(
        archive_summary_path,
        summary_payload,
        encoding="utf-8",
        volatile_field_paths=(("generatedAt",),),
    )

    archive_payload = archive_dashboard_existing if not archive_dashboard_changed and archive_dashboard_existing is not None else payload
    archive_summary = archive_summary_existing if not archive_summary_changed and archive_summary_existing is not None else summary_payload
    report_date_label = str(archive_payload.get("meta", {}).get("reportDateLabel") or report_date.isoformat())
    archive_index = read_json_file(index_path, encoding="utf-8") or {}
    existing_months = archive_index.get("months", [])
    month_entries = [item for item in existing_months if isinstance(item, dict) and item.get("key") != report_month]
    month_entries.append(
        {
            "key": report_month,
            "year": report_date.year,
            "month": report_date.month,
            "label": f"{report_date.year} 年 {report_date.month} 月",
            "reportDate": report_date.isoformat(),
            "reportDateLabel": report_date_label,
            "dashboardPath": build_docs_data_url(archive_dashboard_path, docs_root=docs_root),
            "summaryPath": build_docs_data_url(archive_summary_path, docs_root=docs_root),
            "generatedAt": archive_summary.get("generatedAt"),
        }
    )
    month_entries.sort(key=lambda item: str(item.get("key") or ""), reverse=True)
    index_payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "latestMonth": report_month,
        "months": month_entries,
    }
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_changed, _ = write_json_if_changed(
        index_path,
        index_payload,
        encoding="utf-8",
        volatile_field_paths=(("generatedAt",),),
    )
    return {
        "monthKey": report_month,
        "dashboardPath": build_docs_data_url(archive_dashboard_path, docs_root=docs_root),
        "summaryPath": build_docs_data_url(archive_summary_path, docs_root=docs_root),
        "indexPath": build_docs_data_url(index_path, docs_root=docs_root),
        "dashboardChanged": archive_dashboard_changed,
        "summaryChanged": archive_summary_changed,
        "indexChanged": index_changed,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build dashboard JSON from leads workbook and append arrival brief.")
    parser.add_argument("--workbook", default=str(LEADS_BOOK), help="Path to the leads workbook")
    parser.add_argument("--arrival-workbook", default=str(ARRIVAL_BOOK), help="Path to the arrival workbook")
    parser.add_argument("--out", default=str(OUT_JSON), help="Output JSON path")
    parser.add_argument("--summary-out", default="", help="Optional output summary JSON path")
    parser.add_argument("--report-date", default="", help="Optional report date override in YYYY-MM-DD or YYYYMMDD")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    leads_path = Path(args.workbook).resolve()
    arrival_path = Path(args.arrival_workbook).resolve()
    report_date_override = None
    if args.report_date:
        report_date_override = coerce_date(args.report_date)
        if report_date_override is None:
            raise ValueError(f"Invalid --report-date value: {args.report_date}")
    payload = build_payload(leads_path, arrival_path, report_date_override=report_date_override)
    out_path = Path(args.out).resolve()
    summary_path = Path(args.summary_out).resolve() if args.summary_out else out_path.with_name(f"{out_path.stem}.summary{out_path.suffix}")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    changed, existing_payload = write_json_if_changed(
        out_path,
        payload,
        encoding="utf-8",
        volatile_field_paths=(("meta", "generatedAt"),),
    )
    if not changed and existing_payload is not None:
        payload = existing_payload
    status = "updated" if changed else "unchanged"
    print(f"dashboard.json {status}: {out_path}")
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    archive_seed_summary = build_run_summary(payload, leads_path, arrival_path, out_path, summary_path, changed)
    archive_info = write_monthly_archive(payload, archive_seed_summary)
    summary_payload = build_run_summary(
        payload,
        leads_path,
        arrival_path,
        out_path,
        summary_path,
        changed,
        archive_info=archive_info,
    )
    summary_changed, _ = write_json_if_changed(
        summary_path,
        summary_payload,
        encoding="utf-8",
        volatile_field_paths=(("generatedAt",),),
    )
    summary_status = "updated" if summary_changed else "unchanged"
    print(f"dashboard.summary.json {summary_status}: {summary_path}")
    archive_status = "updated" if archive_info["dashboardChanged"] or archive_info["summaryChanged"] or archive_info["indexChanged"] else "unchanged"
    print(f"dashboard archive {archive_status}: {archive_info['monthKey']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
